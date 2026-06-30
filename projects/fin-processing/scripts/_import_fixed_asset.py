#!/usr/bin/env python3
import openpyxl, json, os, sys, re
sys.stdout.reconfigure(encoding='utf-8')

BASE = os.path.dirname(os.path.abspath(__file__))
indir = r'media/inbound'
fn = [f for f in os.listdir(indir) if 'HOE' in f and ('b2ac6208' in f or 'зӣҳзӮ№' in f)][0]
fp = os.path.join(indir, fn)
wb = openpyxl.load_workbook(fp, data_only=True)

# е…Ҙеә“еҲ°FBдёӯдё“й—Ёзҡ„fixed_assetзұ»еһӢ
fb_fp = os.path.join(BASE, "fb_graph.json")
fb = json.load(open(fb_fp, 'r', encoding='utf-8'))
es = fb.get('entities', [])
existing_ids = set(e.get('id','') for e in es)

# жё…зҗҶж—§иө„дә§ж•°жҚ?for prefix in ['FA_DEPT_', 'FA_ITEM_', 'FA_SUMMARY']:
    es[:] = [e for e in es if not e.get('id','').startswith(prefix)]
existing_ids = set(e.get('id','') for e in es)

# 1. жҖ»иЎЁе…Ҙеә“
summary_sheets = {
    'е®ўжҲҝйғ?: {'sheet': 'е®ўжҲҝйғ?, 'total': 0},
    'йӨҗйҘ®йғ?: {'sheet': 'йӨҗйҘ®йғ?, 'total': 0},
    'еүҚеҺ…йғ?: {'sheet': 'еүҚеҺ…йғ?, 'total': 0},
    'е·ҘзЁӢйғ?: {'sheet': 'е·ҘзЁӢйғ?, 'total': 0},
    'е•ҶеҠЎеҸ‘еұ•йғ?: {'sheet': 'е•ҶеҠЎеҸ‘еұ•йғ?, 'total': 0},
    'дәәдәӢйғ?: {'sheet': 'дәәдәӢйғ?, 'total': 0},
    'иҙўеҠЎйғ?: {'sheet': 'иҙўеҠЎйғ?, 'total': 0},
    'IT': {'sheet': 'IT', 'total': 0},
    'иЎҢж”ҝеҠ?: {'sheet': 'иЎҢж”ҝеҠ?, 'total': 0},
    'е®үе…Ёдҝқйҡңйғ?: {'sheet': 'е®үе…Ёдҝқйҡңйғ?, 'total': 0},
}

all_items = []
grand_total = 0

for dept_name, info in summary_sheets.items():
    ws = wb[info['sheet']]
    dept_items = []
    dept_total = 0
    
    for r in range(4, ws.max_row+1):
        name = str(ws.cell(r, 2).value or '').strip()
        if not name or name == 'еҗҲи®Ў' or len(name) < 2: continue
        qty = ws.cell(r, 5).value or 0
        price = ws.cell(r, 6).value or 0
        amount = ws.cell(r, 7).value or 0
        try:
            qty = int(float(str(qty)))
            price = float(str(price).replace(',',''))
            amount = float(str(amount).replace(',',''))
        except:
            continue
        if qty == 0 and amount == 0: continue
        
        brand = str(ws.cell(r, 3).value or '').strip()
        model = str(ws.cell(r, 4).value or '').strip()
        location = str(ws.cell(r, 11).value or '').strip()
        hoe_ref = str(ws.cell(r, 14).value or '').strip()
        
        dept_items.append({
            'name': name, 'brand': brand, 'model': model,
            'qty': qty, 'unit_price': round(price,2),
            'amount': round(amount,2),
            'location': location, 'hoe_ref': hoe_ref
        })
        dept_total += amount
    
    info['total'] = round(dept_total, 2)
    info['count'] = len(dept_items)
    grand_total += dept_total
    all_items.extend(dept_items)

# йғЁй—ЁжұҮжҖ»иҠӮзӮ?for dept_name, info in summary_sheets.items():
    eid = f"FA_DEPT_{info['sheet']}"
    if eid not in existing_ids:
        es.append({
            "id": eid, "type": "fixed_asset_dept",
            "label": f"{dept_name}еӣәе®ҡиө„дә§",
            "department": dept_name,
            "asset_count": info.get('count', 0),
            "total_amount": info['total'],
            "source": "HOEеӣәе®ҡиө„дә§зӣҳзӮ№202008"
        })
        existing_ids.add(eid)

# е…ЁеҹҹжұҮжҖ»иҠӮзӮ?grand_total = round(grand_total, 2)
total_items = len(all_items)
fa_summary = {
    "id": "FA_SUMMARY_202008",
    "type": "fixed_asset_summary",
    "label": "2020е№?жңҲй…’еә—еӣәе®ҡиө„дә§зӣҳзӮ№жұҮжҖ?,
    "total_amount": grand_total,
    "total_items": total_items,
    "department_count": len(summary_sheets),
    "date": "2020-08",
    "source": "HOEеӣәе®ҡиө„дә§зӣҳзӮ№202008.xlsx"
}
if fa_summary['id'] not in existing_ids:
    es.append(fa_summary)
    existing_ids.add(fa_summary['id'])

# еҗ„жқЎиө„дә§жҳҺз»Ҷ
for i, item in enumerate(all_items):
    eid = f"FA_ITEM_{i+1:05d}"
    if eid not in existing_ids:
        es.append({
            "id": eid, "type": "fixed_asset_item",
            "label": item['name'][:40],
            "brand": item['brand'], "model": item['model'][:40],
            "qty": item['qty'], "unit_price": item['unit_price'],
            "amount": item['amount'], "location": item['location'][:30],
            "hoe_ref": item['hoe_ref'][:30],
        })
        existing_ids.add(eid)

# зүҲжң¬
for v in [e for e in es if e.get('type')=='version' and 'FIN_VER' in e.get('id','')]:
    es.remove(v)
es.append({'id':'FIN_VER_v5_19_FIXED_ASSET','type':'version','label':'FIN v5.19 - HOEеӣәе®ҡиө„дә§зӣҳзӮ№е…ЁйҮҸе…Ҙеә“'})

fb['entities'] = es
json.dump(fb, open(fb_fp, 'w', encoding='utf-8'), ensure_ascii=False, indent=2)

print(f'=== еӣәе®ҡиө„дә§зӣҳзӮ№е…Ҙеә“е®ҢжҲҗ ===')
print(f'жҖ»иө„дә? ВҘ{grand_total:,.2f}')
print(f'жҖ»жқЎзӣ? {total_items}жқ?)
print(f'FB-HOEжҖ»е®һдҪ? {len(es)}')
print(f'\nеҗ„йғЁй—ЁеҲҶеё?')
for dept_name, info in sorted(summary_sheets.items(), key=lambda x: -x[1]['total']):
    print(f'  {dept_name:8s}: ВҘ{info["total"]:>8,.2f}  ({info.get("count",0)}йЎ?')
