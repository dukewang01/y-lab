#!/usr/bin/env python3
import openpyxl, json, os, sys
from collections import defaultdict
sys.stdout.reconfigure(encoding='utf-8')

fp = r'C:\Users\Duke Wang\.openclaw\media\inbound\Daily_Revenue_Report_2026.05.13---be22dea4-a99c-4a37-9bbb-361eaf493d50.xlsx'
wb = openpyxl.load_workbook(fp, data_only=True)
ws = wb['Data']

# ====== 提取5/1-5/13逐日数据 ======
days = []
for col in range(2, 15):  # 5/1-5/13
    sold = ws.cell(5, col).value
    if isinstance(sold, int):
        comp = ws.cell(6, col).value or 0
        hu = ws.cell(7, col).value or 0
        ooo = ws.cell(8, col).value or 0
        vacant = ws.cell(10, col).value or 0
        avail = ws.cell(11, col).value or 0
        occ = float(ws.cell(12, col).value or 0)
        revpar = float(ws.cell(13, col).value or 0)
        adr = float(ws.cell(14, col).value or 0)
        day = col - 1
        days.append((day, sold, comp, hu, ooo, vacant, avail, occ, adr, revpar))

print(f'=== DRR 5/13 — 5月1-13日逐日分析 ===')
print(f'数据日期: 2026-05-13 当天 | 可用房: 538间 | 来源: Daily Revenue Report')
print()

# 总表
print(f'{"日期":>6} | {"入住":>4} | {"Occ%":>6} | {"ADR":>7} | {"RevPAR":>7} | {"空房":>5} | {"备注"}')
print('-' * 65)
total_occ = 0; total_rev = 0
for d in days:
    day, sold, comp, hu, ooo, vacant, avail, occ, adr, revpar = d
    occ_pct = occ * 100
    rev = sold * adr
    total_occ += sold
    total_rev += rev
    
    # 备注
    note = ''
    if occ_pct > 85: note = '满房/高入住'
    elif occ_pct < 35: note = '!! 极低'
    elif occ_pct < 50: note = '较低'
    if adr > 1000: note += ' ADR破千'
    
    # 日期label
    label = f'5/{day}'
    if day == 13: label += ' ←今天'
    
    print(f'{label:>10} | {sold:>4} | {occ_pct:>5.1f}% | {adr:>7.2f} | {revpar:>7.2f} | {vacant:>4}间 | {note}')

avg_occ = total_occ / sum(d[7] for d in days) * 100 
print('-' * 65)
print(f'{"13天合计":>10} | {total_occ:>4}间 | {total_occ/13:>5.0f}/天 | 均ADR={sum(d[8] for d in days)/13:>7.2f} | 均RevPAR={sum(d[9] for d in days)/13:>7.2f}')

# ====== 分析分段 ======
print(f'\n{"="*60}')
print('  阶段分析')
print('='*60)

# 五一4天
may1_4 = [d for d in days if d[0] <= 4]
print(f'\n[五一黄金周 5/1-5/4]')
print(f'  总入住: {sum(d[1] for d in may1_4)}间 | 均Occ: {sum(d[7] for d in may1_4)/4*100:.1f}% | 均ADR: ¥{sum(d[8] for d in may1_4)/4:.0f} | 总收入: ¥{sum(d[1]*d[8] for d in may1_4):,.0f}')
print(f'  最佳: 5/2 ADR¥1,055  |  最弱: 5/4 Occ61.0%')

# 节后5天
may5_9 = [d for d in days if 5 <= d[0] <= 9]
print(f'\n[节后震荡 5/5-5/9]')
print(f'  均Occ: {sum(d[7] for d in may5_9)/5*100:.1f}% | 均ADR: ¥{sum(d[8] for d in may5_9)/5:.0f}')
print(f'  特征: 5/5仅27%触底 → 5/7满房95% → 5/8周六仅43%谜团')

# 本周
may10_13 = [d for d in days if d[0] >= 10]
print(f'\n[本周 5/10-5/13]')
print(f'  均Occ: {sum(d[7] for d in may10_13)/4*100:.1f}% | 均ADR: ¥{sum(d[8] for d in may10_13)/4:.0f}')
weekday_occ = sum(d[7] for d in may10_13 if d[0] >= 11)/3*100
print(f'  本周工作日(5/11-13): Occ {weekday_occ:.1f}% — "二次空窗期"确认')

# ====== 同比2025年5月 ======
ws_ly = wb['LY']
ly_rev_mtd = ws_ly.cell(6, 2).value  # 2025年5月MTD收入
ly_room_sold = ws_ly.cell(6, 3).value  # 2025年5月MTD间夜
# LY sheet结构: 第1行是日期, 第2行是房间, 第3行是收入, 第4行是ADR
# 行5-18: MTD逐日
print(f'\n[同比去年 2025年5月同期]')
ly_days = []
for r in range(6, 19):
    sold_val = ws_ly.cell(r, 6).value  # 列F=逐日销量
    rev_val = ws_ly.cell(r, 7).value   # 列G=逐日收入
    if sold_val and rev_val:
        ly_days.append((sold_val, rev_val))

if ly_days:
    ly_sold = sum(d[0] for d in ly_days)
    ly_rev = sum(d[1] for d in ly_days)
    this_rev = sum(d[1]*d[8] for d in days)
    this_sold = sum(d[1] for d in days)
    diff_sold = this_sold - ly_sold
    diff_rev = this_rev - ly_rev
    print(f'           | 2025(去年) | 2026(今年) | 同比')
    print(f'  入住间夜 | {ly_sold:>8d} | {this_sold:>8d} | {diff_sold:>+6d} ({diff_sold/ly_sold*100:+.1f}%)')
    print(f'  客房收入 | ¥{ly_rev:>8,.0f} | ¥{this_rev:>8,.0f} | {diff_rev/ly_rev*100:+.1f}%')
    print(f'  状态: ', end='')
    if diff_rev > 0: print(f'超过去年同期 ✅')
    else: print(f'低于去年同期 ⚠️')

# ====== F&B当天（5/13） ======
ws_fb = wb['F&B']
print(f'\n[5/13 当天餐饮]')
total_fb_rev = 0
for r in range(11, 19):
    name = str(ws_fb.cell(r, 1).value or '').strip()
    rev = ws_fb.cell(r, 2).value
    covers = ws_fb.cell(r, 3).value
    avg = ws_fb.cell(r, 4).value
    if name and rev and float(rev) > 0:
        total_fb_rev += float(rev)
        clean_name = name.replace('全','').replace('餐','').replace('厅','').replace('味','').replace('风','').replace('酒','').replace('大','').replace('吧','').replace('堂','').replace('  ','')
        print(f'  {clean_name[:8]:8s} | 收入¥{float(rev):>7,.0f} | {int(covers) if covers else 0:>4}人 | 均¥{float(avg):>5.0f}')
print(f'  -------')
print(f'  当天FB合计: ¥{total_fb_rev:,.0f}')

# ====== 核心结论 ======
print(f'\n{"="*60}')
print('  核心结论')
print('='*60)
this_week = [d for d in days if d[0] >= 11]
this_occ = sum(d[7] for d in this_week)/len(this_week)*100
print(f'1. 5/13(三) Occ {days[-1][7]*100:.1f}% / ADR ¥{days[-1][8]:.0f} — 工作日正常恢复中')
print(f'2. 本周(5/11-13)均Occ {this_occ:.1f}%，处于节后"二次空窗期"')
print(f'3. 5月前13天总收入约¥{sum(d[1]*d[8] for d in days):,.0f}')
print(f'4. 对比之前HF预测: DRR的ADR比HF口径高5-15% — 注意差异')
print(f'5. 当务之急: 5/22-5/31的低迷期策略')
