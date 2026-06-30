"""
精确定位2026年虫害投诉
"""
import openpyxl, os, re, json
from datetime import datetime
from collections import Counter

BASE = 'C:\\Users\\Duke Wang\\.openclaw\\workspace\\knowledge_center'
LOG_DIR = os.path.join(BASE, 'log_cases')
files_2026 = [f for f in os.listdir(LOG_DIR) if '2026' in f and f.endswith('.xlsx')]

real_pest = []
for fname in files_2026:
    fp = os.path.join(LOG_DIR, fname)
    wb = openpyxl.load_workbook(fp, read_only=True, data_only=True)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2: continue
        for row in rows[1:]:
            if not row or not any(row): continue
            text_fields = [str(c) for c in row if c and isinstance(c, (str, datetime))]
            full = ' '.join(text_fields)
            
            pest_found = []
            checks = [
                ('蟑螂', ['蟑螂','cockroach','cockroach']),
                ('老鼠/老鼠屎', ['老鼠','鼠粪','鼠屎','鼠','mouse','rat']),
                ('蚂蚁', ['蚂蚁','ant']),
                ('蚊子', ['蚊子','mosquito']),
                ('苍蝇', ['苍蝇','fly']),
                ('虫子/虫害', ['虫子','虫爬','飞虫','虫害','虫咬']),
            ]
            for label, kws in checks:
                if any(kw.lower() in full.lower() for kw in kws):
                    pest_found.append(label)
            
            if pest_found:
                case_id = str(row[0])[:10] if row[0] else '?'
                room = str(row[5]).strip()[:10] if len(row) > 5 and row[5] else ''
                date_val = row[2] if len(row) > 2 else ''
                channel = str(row[8])[:20] if len(row) > 8 and row[8] else ''
                amt_raw = row[10] if len(row) > 10 else 0
                desc = str(row[11])[:300] if len(row) > 11 and row[11] else ''
                desc2 = str(row[12])[:300] if len(row) > 12 and row[12] else ''
                handler = str(row[1])[:15] if len(row) > 1 and row[1] else ''
                
                amount = 0
                if amt_raw:
                    try: amount = float(amt_raw)
                    except:
                        try: amount = float(str(amt_raw).replace(',','').replace('¥','').replace('楼',''))
                        except: pass
                
                day_str = ''
                if isinstance(date_val, datetime):
                    day_str = date_val.strftime('%Y-%m-%d')
                elif isinstance(date_val, str):
                    m = re.search(r'(\d{4}-\d{2}-\d{2})', date_val)
                    if m: day_str = m.group(1)
                
                real_pest.append({
                    'id': case_id,
                    'date': day_str,
                    'room': room,
                    'channel': channel,
                    'amount': amount,
                    'handler': handler,
                    'type': pest_found,
                    'desc': (desc + ' ' + desc2).strip()[:250],
                })
    wb.close()

print(f'===== 2026年真实虫害投诉 =====')
print(f'共 {len(real_pest)} 起')

if real_pest:
    # 排重（同一个案例ID可能出现在多个sheet里）
    seen_ids = set()
    unique_cases = []
    for c in real_pest:
        if c['id'] not in seen_ids:
            seen_ids.add(c['id'])
            unique_cases.append(c)
    
    print(f'去重后: {len(unique_cases)} 起')
    
    # 类型统计
    type_stats = Counter()
    for c in unique_cases:
        for t in c['type']:
            type_stats[t] += 1
    
    print(f'\n虫害类型:')
    for t, cnt in type_stats.most_common():
        print(f'  {t}: {cnt} 起')
    
    # 月份
    print(f'\n月度分布:')
    months = Counter()
    for c in unique_cases:
        if c['date']:
            m = c['date'][5:7]
            months[int(m)] += 1
    for m in range(1,13):
        if months[m]:
            print(f'  {m}月: {months[m]} 起')
    
    # 赔偿
    amounts = [c['amount'] for c in unique_cases if c['amount'] > 0]
    if amounts:
        print(f'\n赔偿: {len(amounts)} 起, 共¥{sum(amounts):,.0f}, 平均¥{sum(amounts)/len(amounts):,.0f}')
    
    # 处理人
    handlers = Counter(c['handler'] for c in unique_cases if c['handler'])
    if handlers:
        print(f'\n处理人:')
        for h, cnt in handlers.most_common():
            print(f'  {h}: {cnt} 起')
    
    # 详细案例
    print(f'\n详细案例:')
    for c in unique_cases:
        room = f'房{c["room"]}' if c['room'] else ''
        amt = f'赔偿¥{c["amount"]:.0f}' if c['amount'] > 0 else ''
        types = ','.join(c['type'])
        print(f'')
        print(f'  #{c["id"]} | {c["date"]} | {room} | {types} {amt}')
        print(f'  {c["desc"][:200]}')
