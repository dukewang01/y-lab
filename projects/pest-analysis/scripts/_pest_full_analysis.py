"""
虫害全维分析 - 跨站数据聚合
GSM + FSAA + FAQ + Risk + 2026投诉案例
"""
import json, os, sys
sys.stdout.reconfigure(encoding='utf-8')

from collections import Counter, defaultdict
from datetime import datetime

BASE = 'C:\\Users\\Duke Wang\\.openclaw\\workspace\\knowledge_center'

print('='*65)
print('    苏州希尔顿 · 虫害全维分析报告')
print('    分析日期: 2026-05-07')
print('='*65)
print()

# ====== 1. GSM站虫害知识 ======
print('【一、虫害知识体系（GSM站）】')
with open(os.path.join(BASE, 'gsm_graph.json'), 'r', encoding='utf-8-sig') as f:
    gsm = json.load(f)

# 找虫害相关的全部实体（不限定ID，搜全部属性）
all_text_gsm = ''
pest_knowledge = []
for e in gsm['entities']:
    eid = e.get('id','')
    props = str(e.get('properties',{}))
    text = eid + ' ' + props
    all_text_gsm += text + '\n'
    for kw in ['pest','虫害','蟑螂','老鼠','虫','insect']:
        if kw.lower() in text.lower():
            pest_knowledge.append(e)
            break

print(f'  虫害相关节点: {len(pest_knowledge)}')

# SOP和场景
sops = [e for e in pest_knowledge if e.get('type') == 'sop']
scenes = [e for e in pest_knowledge if e.get('type') == 'gsm_scene']
concepts = [e for e in pest_knowledge if e.get('type') == 'concept']
print(f'  SOP流程: {len(sops)}')
for s in sops:
    print(f'    📋 {s["id"]}')
print(f'  专项场景: {len(scenes)}')
for s in scenes:
    print(f'    🎯 {s["id"]}')
print()

# ====== 2. FSAA站虫害体系 ======
print('【二、食品安全与虫害控制（FSAA站）】')
with open(os.path.join(BASE, 'fsaa_graph.json'), 'r', encoding='utf-8-sig') as f:
    fsaa = json.load(f)

pest_fsaa = []
for e in fsaa['entities']:
    eid = e.get('id','')
    props = str(e.get('properties',{}))
    text = eid + ' ' + props
    for kw in ['pest','虫害','虫','蟑螂','老鼠','苍蝇','pesticide','ipm','integrated pest','消杀','除虫']:
        if kw.lower() in text.lower():
            pest_fsaa.append(e)
            break

print(f'  虫害相关节点: {len(pest_fsaa)}')

# 按类型统计
type_counts = Counter()
for e in pest_fsaa:
    type_counts[e.get('type','?')] += 1
for t, c in type_counts.most_common():
    print(f'    {t}: {c}')
print()

# ====== 3. FAQ站虫害知识 ======
print('【三、虫害FAQ知识库】')
with open(os.path.join(BASE, 'faq_graph.json'), 'r', encoding='utf-8-sig') as f:
    faq = json.load(f)

pest_faq = []
for e in faq['entities']:
    props = e.get('properties',{})
    text = str(props.get('question','')) + str(props.get('answer','')) + str(props.get('标签',''))
    for kw in ['虫','蟑螂','老鼠','蚊子','苍蝇','蚂蚁','pest']:
        if kw in text:
            pest_faq.append(e)
            break

print(f'  虫害相关FAQ: {len(pest_faq)} 条')
for e in pest_faq[:8]:
    props = e.get('properties',{})
    q = str(props.get('question',''))[:60]
    a = str(props.get('answer',''))[:80]
    tag = str(props.get('标签',''))[:20]
    print(f'  📖 Q: {q}')
    print(f'     A: {a}')
    if tag: print(f'     标签: {tag}')
print()

# ====== 4. Risk站虫害风险案例 ======
print('【四、虫害风险案例（Risk站）】')
with open(os.path.join(BASE, 'risk_graph.json'), 'r', encoding='utf-8-sig') as f:
    risk = json.load(f)

pest_risk = []
for e in risk['entities']:
    text = e.get('id','') + ' ' + str(e.get('properties',{}))
    for kw in ['pest','虫害','虫','蟑螂','老鼠','苍蝇','蚂蚁']:
        if kw in text:
            pest_risk.append(e)
            break

print(f'  虫害风险案例: {len(pest_risk)}')
for e in pest_risk[:5]:
    props = e.get('properties',{})
    desc = str(props.get('description',''))[:100]
    severity = props.get('severity','')
    print(f'  ⚠️ {e["id"]}: {desc}')
    if severity: print(f'     风险等级: {severity}')
print()

# ====== 5. 2026年投诉案例中的虫害问题 ======
print('【五、2026年虫害投诉案例】')
import openpyxl

LOG_DIR = os.path.join(BASE, 'log_cases')
files_2026 = [f for f in os.listdir(LOG_DIR) if '2026' in f and f.endswith('.xlsx') and '副本' not in f]

pest_cases = []
pest_keywords_total = defaultdict(int)

for fname in files_2026:
    fp = os.path.join(LOG_DIR, fname)
    wb = openpyxl.load_workbook(fp, read_only=True, data_only=True)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2: continue
        for row_idx, row in enumerate(rows[1:], 2):
            if not row or not any(row): continue
            
            case_id = str(row[0])[:10] if row[0] else f'{fname[:5]}_Row{row_idx}'
            date_val = row[2] if len(row) > 2 else ''
            time_str = str(row[3])[:10] if len(row) > 3 and row[3] else ''
            guest = str(row[4])[:20] if len(row) > 4 and row[4] else ''
            room = str(row[5])[:10] if len(row) > 5 and row[5] else ''
            channel = str(row[8])[:20] if len(row) > 8 and row[8] else ''
            amount_raw = row[10] if len(row) > 10 else 0
            desc = str(row[11])[:500] if len(row) > 11 and row[11] else ''
            desc2 = str(row[12])[:500] if len(row) > 12 and row[12] else ''
            full = desc + ' ' + desc2
            
            # 虫害关键词检测
            pest_words = ['蟑螂','cockroach','老鼠','鼠','mouse','rat','蚂蚁','ant',
                         '蚊子','mosquito','苍蝇','fly','虫子','worm','飞虫','小虫',
                         '虫卵','虫咬','虫爬','bedbug','臭虫','跳蚤','flea','蜈蚣']
            
            matched = [kw for kw in pest_words if kw.lower() in full.lower()]
            if matched:
                amount = 0
                if amount_raw:
                    try:
                        amount = float(amount_raw)
                    except:
                        try:
                            amount = float(str(amount_raw).replace(',','').replace('¥','').replace('楼',''))
                        except:
                            pass
                
                # 解析日期
                day_str = ''
                if isinstance(date_val, datetime):
                    day_str = date_val.strftime('%m/%d')
                elif isinstance(date_val, str):
                    import re
                    m = re.search(r'(\d{4})-(\d{2})-(\d{2})', date_val)
                    if m: day_str = f'{int(m.group(2))}/{int(m.group(3))}'
                
                pest_cases.append({
                    'id': case_id,
                    'date': day_str,
                    'time': str(date_val)[:16],
                    'room': room,
                    'guest': guest,
                    'channel': channel,
                    'amount': amount,
                    'desc': full[:200],
                    'keywords': matched,
                    'file': fname[:15]
                })
                
                for kw in matched:
                    pest_keywords_total[kw] += 1
    wb.close()

print(f'  2026年虫害投诉: {len(pest_cases)} 起')

if pest_cases:
    # 按月份
    month_pest = Counter()
    for c in pest_cases:
        if '/' in c['date']:
            month_pest[c['date'].split('/')[0]] += 1
    print(f'  月度分布:')
    for m in ['1','2','3','4']:
        print(f'    {m}月: {month_pest.get(m,0)} 起')
    
    # 虫害类型
    print(f'  虫害类型:')
    for kw, cnt in sorted(pest_keywords_total.items(), key=lambda x: -x[1]):
        bar = '▇' * cnt
        print(f'    {kw:<8s}: {cnt} {bar}')
    
    # 赔偿
    amounts = [c['amount'] for c in pest_cases if c['amount'] > 0]
    if amounts:
        print(f'  涉及赔偿: {len(amounts)} 起')
        print(f'  总赔偿: ¥{sum(amounts):,.0f}')
        print(f'  平均: ¥{sum(amounts)/len(amounts):,.0f}')
    
    # 详细案例
    print(f'  案例详情:')
    for c in pest_cases[:8]:
        room_info = f'房{c["room"]}' if c['room'] else ''
        amt = f'赔偿¥{c["amount"]:.0f}' if c['amount'] > 0 else ''
        desc_short = c['desc'][:80]
        print(f'    #{c["id"]} {c["date"]} {room_info} {amt}: {desc_short}')

# ====== 6. 跨站关系 ======
print(f'\n【六、虫害知识跨站联动】')
# GSM -> FSAA 关系
gsm_rels = gsm.get('relations', [])
fsaa_ids = {e['id'] for e in fsaa['entities']}
gsm_to_fsaa = [r for r in gsm_rels if r.get('target') in fsaa_ids or r.get('source') in fsaa_ids]
print(f'  GSM ↔ FSAA 跨站关系: {len(gsm_to_fsaa)} 条')

# FAQ标签
print(f'  虫害FAQ标签:')
for e in pest_faq:
    tag = e.get('properties',{}).get('标签','')
    if tag:
        print(f'    {tag}')

print(f'\n{"="*65}')
print(f'分析完毕')
print(f'{"="*65}')
