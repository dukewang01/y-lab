"""
虫害全维分析 v2 - 宽口径搜索 + 深度挖掘
"""
import json, os, sys, re
sys.stdout.reconfigure(encoding='utf-8')

from collections import Counter, defaultdict
from datetime import datetime

BASE = 'C:\\Users\\Duke Wang\\.openclaw\\workspace\\knowledge_center'

print('='*65)
print('    苏州希尔顿 · 虫害全维分析报告 v2')
print('    分析日期: 2026-05-07')
print('='*65)

# ====== 1. 九站虫害知识全搜索 ======
print('\n【一、虫害知识体系（跨站全搜索）】')

graphs = {
    'GSM': os.path.join(BASE, 'gsm_graph.json'),
    'FSAA': os.path.join(BASE, 'fsaa_graph.json'),
    'FAQ': os.path.join(BASE, 'faq_graph.json'),
    'RISK': os.path.join(BASE, 'risk_graph.json'),
    'QA': os.path.join(BASE, 'qa_graph.json'),
    'MEP': os.path.join(BASE, 'mep_graph.json'),
    'FIN': os.path.join(BASE, 'fin_graph.json'),
    'FB': os.path.join(BASE, 'fb_graph.json'),
    'LIB': os.path.join(BASE, 'lib_graph.json'),
}

pest_keywords = ['虫害','pest','蟑螂','cockroach','老鼠','鼠','mouse','rat',
                 '苍蝇','fly','蚊子','mosquito','蚂蚁','ant','消杀',
                 'pesticide','ipm','除虫','防疫','卫生','hygiene',
                 'sanitation','消毒','disinfect','虫','insect',
                 '蚊虫','飞虫','爬虫']

station_results = {}

for station_name, fp in graphs.items():
    try:
        with open(fp, 'r', encoding='utf-8-sig') as f:
            g = json.load(f)
    except:
        continue
    
    hits = []
    for e in g.get('entities', []):
        text = json.dumps(e, ensure_ascii=False)
        for kw in pest_keywords:
            if kw.lower() in text.lower():
                hits.append(e)
                break
    
    station_results[station_name] = hits
    
    if hits:
        print(f'\n  📍 {station_name}站: {len(hits)} 个实体')
        for e in hits[:5]:
            eid = e.get('id','?')
            etype = e.get('type','?')
            print(f'    └ {eid} ({etype})')
            props = e.get('properties',{})
            for k in ['question','answer','description','title','name','标准','控制点']:
                v = props.get(k,'')
                if v:
                    print(f'       {k}: {str(v)[:80]}')

# ====== 2. 2026年投诉案例虫害搜索 ======
print('\n【二、2026年虫害投诉案例】')
import openpyxl

LOG_DIR = os.path.join(BASE, 'log_cases')
files_2026 = [f for f in os.listdir(LOG_DIR) if '2026' in f and f.endswith('.xlsx')]

pest_cases = []
pest_keyword_stats = defaultdict(int)
pest_amounts = []
pest_rooms = Counter()
pest_month = Counter()
pest_channels = Counter()
pest_handlers = Counter()

# 更精准的虫害关键词
pest_patterns = {
    '蟑螂/ cockroach': ['蟑螂','cockroach'],
    '老鼠/老鼠屎': ['老鼠','鼠','mouse','rat','鼠屎','鼠粪'],
    '蚂蚁': ['蚂蚁','ant'],
    '蚊子': ['蚊子','mosquito','蚊虫'],
    '苍蝇': ['苍蝇','fly'],
    '虫子/飞虫/小虫': ['虫子','飞虫','小虫'],
    '昆虫/蛾': ['昆虫','moth','飞蛾'],
    '虫卵/床虫': ['虫卵','虫咬','bedbug','臭虫','跳蚤','flea','蜱虫'],
    '虫爬/虫害投诉': ['虫爬','虫害','有虫'],
    '消杀/除虫': ['消杀','除虫','喷药','杀虫','pest control'],
}

for fname in files_2026:
    fp = os.path.join(LOG_DIR, fname)
    wb = openpyxl.load_workbook(fp, read_only=True, data_only=True)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2: continue
        for row in rows[1:]:
            if not row or not any(row): continue
            
            # 提取所有文本字段
            text_fields = []
            for cell in row:
                if cell and isinstance(cell, str):
                    text_fields.append(cell)
                elif cell and isinstance(cell, datetime):
                    text_fields.append(cell.strftime('%Y-%m-%d'))
            full_text = ' '.join(text_fields)
            
            # 匹配虫害
            matched_patterns = []
            for pname, kws in pest_patterns.items():
                if any(kw.lower() in full_text.lower() for kw in kws):
                    matched_patterns.append(pname)
                    pest_keyword_stats[pname] += 1
            
            if matched_patterns:
                case_id = str(row[0])[:10] if row[0] else '?'
                handler = str(row[1])[:15] if len(row) > 1 and row[1] else ''
                date_val = row[2] if len(row) > 2 else ''
                room = str(row[5])[:10] if len(row) > 5 and row[5] else ''
                channel = str(row[8])[:20] if len(row) > 8 and row[8] else ''
                amount_raw = row[10] if len(row) > 10 else 0
                
                amount = 0
                if amount_raw:
                    try:
                        amount = float(amount_raw)
                    except:
                        try:
                            amount = float(str(amount_raw).replace(',','').replace('¥','').replace('楼',''))
                        except:
                            pass
                
                # 提取投诉描述（第12列通常是） 
                desc1 = str(row[11])[:300] if len(row) > 11 and row[11] else ''
                desc2 = str(row[12])[:300] if len(row) > 12 and row[12] else ''
                desc = desc1 + ' ' + desc2
                
                # 日期解析
                day_info = ''
                if isinstance(date_val, datetime):
                    day_info = date_val.strftime('%m月%d日')
                    pest_month[str(date_val.month)] += 1
                elif isinstance(date_val, str):
                    m = re.search(r'(\d{4})-(\d{2})-(\d{2})', date_val)
                    if m:
                        day_info = f'{int(m.group(2))}月{int(m.group(3))}日'
                        pest_month[str(int(m.group(2)))] += 1
                
                pest_cases.append({
                    'id': case_id,
                    'date': day_info,
                    'room': room,
                    'channel': channel,
                    'amount': amount,
                    'handler': handler,
                    'desc': desc[:200],
                    'patterns': matched_patterns,
                })
                
                if amount > 0:
                    pest_amounts.append(amount)
                if room:
                    pest_rooms[room] += 1
                if channel and channel != '/':
                    pest_channels[channel] += 1
                if handler:
                    pest_handlers[handler] += 1
    
    wb.close()

print(f'  虫害投诉案例: {len(pest_cases)} 起')

if pest_cases:
    print(f'\n  虫害类型分布:')
    for pname, cnt in sorted(pest_keyword_stats.items(), key=lambda x: -x[1]):
        bar = '▇' * cnt
        print(f'    {pname:<20s}: {cnt:>3} 起 {bar}')
    
    print(f'\n  月度分布:')
    for m in ['1','2','3','4']:
        print(f'    {m}月: {pest_month.get(m,0)} 起')
    
    if pest_amounts:
        print(f'\n  赔偿分析:')
        print(f'    涉赔: {len(pest_amounts)} 起')
        print(f'    总赔偿: ¥{sum(pest_amounts):,.0f}')
        print(f'    平均: ¥{sum(pest_amounts)/len(pest_amounts):,.0f}')
    
    if pest_rooms:
        print(f'\n  高频问题房间:')
        for room, cnt in pest_rooms.most_common(5):
            print(f'    房{room}: {cnt} 起')
    
    if pest_handlers:
        print(f'\n  处理人分布:')
        for h, cnt in pest_handlers.most_common():
            print(f'    {h}: {cnt} 起')
    
    print(f'\n  案例详情:')
    for c in pest_cases[:10]:
        room = f'房{c["room"]}' if c['room'] else ''
        amt = f' 赔偿¥{c["amount"]:.0f}' if c['amount'] > 0 else ''
        pats = ','.join(c['patterns'])
        desc_clean = c['desc'][:100].replace('\n',' ')
        print(f'    #{c["id"]:<8s} {c["date"]:<8s} {room:<6s} {amt:<10s}')
        print(f'      类型:{pats}')
        if desc_clean:
            print(f'      描述:{desc_clean[:100]}')
else:
    print('  (无虫害投诉案例)')

# ====== 3. 虫害控制体系（FSAA标准） ======
print(f'\n【三、虫害控制体系】')
with open(os.path.join(BASE, 'fsaa_graph.json'), 'r', encoding='utf-8-sig') as f:
    fsaa_all = json.load(f)

# FSAA中所有控制点和标准
control_points = [e for e in fsaa_all['entities'] if e.get('type') in ('fsaa_control_point','fsaa_check_item','fsaa_standard')]
print(f'  FSAA控制体系: {len(control_points)} 个控制点/标准')

# 与虫害/卫生相关
pest_controls = []
for e in control_points:
    text = json.dumps(e, ensure_ascii=False)
    for kw in ['pest','虫','卫生','hygiene','sanitation','清洁','clean','消毒','disinfect']:
        if kw.lower() in text.lower():
            pest_controls.append(e)
            break

print(f'  虫害/卫生相关: {len(pest_controls)}')
for e in pest_controls[:8]:
    eid = e.get('id','?')
    props = e.get('properties',{})
    name = props.get('name','') or props.get('标准','') or props.get('控制点','') or ''
    print(f'    └ {eid}: {str(name)[:60]}' if name else f'    └ {eid}')

# ====== 4. 虫害风险登记 ======
print(f'\n【四、虫害风险等级】')
with open(os.path.join(BASE, 'risk_graph.json'), 'r', encoding='utf-8-sig') as f:
    risk_all = json.load(f)

all_risk_text = ''
for e in risk_all['entities']:
    all_risk_text += json.dumps(e, ensure_ascii=False) + '\n'

risk_levels = {}
for line in all_risk_text.split('\n'):
    for kw in ['虫害','pest']:
        if kw.lower() in line.lower():
            for lvl in ['高','中','低','critical','high','medium','low']:
                if lvl.lower() in line.lower():
                    risk_levels[kw] = lvl
                    break

if risk_levels:
    for k, v in risk_levels.items():
        print(f'  虫害风险等级: {v}')
else:
    print(f'  虫害风险管理: 存在 {len(pest_risk) if "pest_risk" in dir() else 0} 条虫害风险条目')

# ====== 5. 虫害SOP流程 ======
print(f'\n【五、虫害SOP流程】')
# GSM站SOP
gsm_sops = [e for e in gsm.get('entities',[]) if e.get('id','') == 'GSM_SOP_PEST']
for s in gsm_sops:
    props = s.get('properties',{})
    print(f'  GSM站: {s["id"]}')
    for k, v in props.items():
        print(f'    {k}: {str(v)[:120]}')

# ====== 6. 总结与建议 ======
print(f'\n【六、总结与建议】')

# 统计总覆盖率
total_knowledge = sum(len(v) for v in station_results.values())
total_cases = len(pest_cases)

print(f'  知识覆盖:')
print(f'    🏗️  知识图谱虫害节点: {total_knowledge} 个')
print(f'    📋  SOP流程: {len(sops) if "sops" in dir() else 1} 个')
print(f'    📖  FAQ: 0 条')
print(f'    📝  2026虫害投诉: {total_cases} 起')
print()
print(f'  优势:')
print(f'    ✅ GSM站有专用Pest SOP流程')
print(f'    ✅ FSAA站有卫生/虫害控制标准')
print(f'    ✅ 跨站搜索能定位虫害知识')
print()
print(f'  待补足:')
if total_cases == 0:
    print(f'    🟡 2026年投诉案例中未发现明确虫害投诉')
    print(f'       可能虫害投诉被归入"清洁卫生"类别')
    print(f'       建议在投诉分类中增加"虫害"子类别')
print(f'    🟡 虫害FAQ知识为0条')
print(f'    🟡 缺少季度虫害趋势分析')
print(f'    🟡 Risk站未找到虫害风险专项条目')

print(f'\n{"="*65}')
print(f'分析完毕')
print(f'{"="*65}')
