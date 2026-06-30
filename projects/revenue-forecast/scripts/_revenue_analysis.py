import json, pdfplumber
from datetime import datetime, date
from collections import defaultdict

# === 1. 读图谱 ===
fin_file = 'C:\\Users\\Duke Wang\\.openclaw\\workspace\\knowledge_center\\fin_graph.json'
with open(fin_file, 'r', encoding='utf-8-sig') as f:
    graph = json.load(f)

ents = graph['entities']
rels = graph['relations']

# === 2. 提取所有日报节点（5月） ===
may_dailies = {}
for e in ents:
    if e.get('type') == 'daily_revenue':
        eid = e.get('id','')
        if '2026_05' in eid:
            props = e.get('properties', {})
            day = eid.split('_')[-1]  # "05" → "06"?
            may_dailies[eid] = {
                'name': e.get('name',''),
                'revenue': props.get('room_revenue', 0),
                'rooms_sold': props.get('rooms_sold', props.get('room_sold', 0)),
                'adr': props.get('adr', 0),
                'occupancy': props.get('occupancy_pct', props.get('occ_pct', 0)),
                'revpar': props.get('revpar', 0),
                'guest_count': props.get('guest_count', 0),
                'source': props.get('data_type', 'unknown')
            }

print('=== 5月已有日报数据 ===')
total_rev = 0
total_rooms = 0
days_count = 0
days_actual = 0

for eid in sorted(may_dailies.keys()):
    d = may_dailies[eid]
    day = eid.split('_')[-1]
    src = d.get('source','')
    print(f'  5月{day}日: 收入={d["revenue"]}, 售房={d["rooms_sold"]}, ADR={d["adr"]}, 出租率={d["occupancy"]:.2%}')
    if d['revenue'] and float(str(d['revenue']).replace(',','')) > 0:
        rev = float(str(d['revenue']).replace(',',''))
        total_rev += rev
        total_rooms += int(float(str(d['rooms_sold']).replace(',','')))
        days_count += 1
        if 'actual' in src.lower() or 'forecast' not in src.lower():
            days_actual += 1

print(f'\n  累计: 收入={total_rev:.2f}, 售房={total_rooms}间, {days_count}天有数据')
print(f'  其中实际数据: {days_actual}天')

# === 3. 读最新日报（5月6日详细）===
print('\n=== 5月6日详细数据 ===')
xlsx = 'C:\\Users\\Duke Wang\\.openclaw\\knowledge_center\\media\\Daily_Revenue_Report_2026.05.06.xlsx'
import openpyxl
wb = openpyxl.load_workbook(xlsx, data_only=True)
ws = wb['Actual']

def get_val(ws, label, col=5):
    for r in range(1, ws.max_row+1):
        a = str(ws.cell(r, 4).value or '').strip()
        ws2 = wb['F&B']
        if label.lower() in a.lower():
            return ws.cell(r, col).value
    return None

def get_fb_row(ws, outlet_name):
    """从F&B表找对应营业点行"""
    for r in range(1, ws.max_row+1):
        a = str(ws.cell(r, 1).value or '').strip()
        if outlet_name.lower() in a.lower():
            day_rev = ws.cell(r, 2).value  # 当日收入
            day_cov = ws.cell(r, 3).value  # 当日人数
            mtd_rev = ws.cell(r, 7).value  # 月累计收入
            mtd_cov = ws.cell(r, 6).value  # 月累计人数
            ly_mtd_rev = ws.cell(r, 8).value  # 上年当月累计
            budget_mtd_rev = ws.cell(r, 9).value  # 业主月预算
            return {'day_rev': day_rev, 'day_cov': day_cov, 
                    'mtd_rev': mtd_rev, 'mtd_cov': mtd_cov,
                    'ly_mtd_rev': ly_mtd_rev, 'budget_mtd_rev': budget_mtd_rev}
    return None

ws2 = wb['F&B']
outlets = ['BANQUET AND CONFEREN','OPEN','YUXI','BACIO','BEER SOCIETY','YUAN',
           'FOOD STORE','ROOM SERVICE','MINI BAR']
print('营业点详细:')
for o in outlets:
    d = get_fb_row(ws2, o)
    if d:
        print(f'  {o}:')
        print(f'    当日: 收入={d["day_rev"]}, 人数={d["day_cov"]}')
        if d['mtd_rev']:
            vs_ly = (d.get('mtd_rev',0) or 0) - (d.get('ly_mtd_rev',0) or 0)
            vs_budget = (d.get('mtd_rev',0) or 0) - (d.get('budget_mtd_rev',0) or 0)
            print(f'    月累计: 收入={d["mtd_rev"]:.2f}, 人数={d["mtd_cov"]}')
            print(f'    vs上年: {vs_ly:+.2f}, vs预算: {vs_budget:+.2f}')

# Total F&B
print()
d = get_fb_row(ws2, 'TOTAL F&B')
if d:
    print(f'  F&B合计:')
    print(f'    当日: 收入={d["day_rev"]}, 人数={d["day_cov"]}')
    vs_ly = (d.get('mtd_rev',0) or 0) - (d.get('ly_mtd_rev',0) or 0)
    vs_budget = (d.get('mtd_rev',0) or 0) - (d.get('budget_mtd_rev',0) or 0)
    print(f'    月累计: 收入={d["mtd_rev"]:.2f}, 人数={d["mtd_cov"]}')
    print(f'    vs上年: {vs_ly:+.2f}, vs预算: {vs_budget:+.2f}')

# === 4. 读预测PDF ===
print('\n=== 历史与预测 (PDF) ===')
pdf_file = 'C:\\Users\\Duke Wang\\.openclaw\\knowledge_center\\source_files\\FIN\\History_and_Forecast_5.6.pdf'
with pdfplumber.open(pdf_file) as pdf:
    for i, page in enumerate(pdf.pages):
        text = page.extract_text()
        if text:
            lines = text.strip().split('\n')
            print(f'  [第{i+1}页] {lines[0][:100]}')
            for line in lines[1:6]:
                print(f'    {line[:120]}')
