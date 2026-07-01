#!/usr/bin/env python3
"""RevPAR efficiency deep dive for Hotel-A."""
import openpyxl, math

path = r'C:\Users\Y\.openclaw\workspace\media\archived\GCM_YTD.xlsx'
wb = openpyxl.load_workbook(path, data_only=True)
ws = wb['Export']

markets = {}
hotels = {}
current_market = None
for r in range(2, ws.max_row + 1):
    c1 = ws.cell(r, 1).value; c2 = ws.cell(r, 2).value
    c3 = ws.cell(r, 3).value; c4 = ws.cell(r, 4).value
    c5 = ws.cell(r, 5).value; c6 = ws.cell(r, 6).value; c7 = ws.cell(r, 7).value
    if c1 is not None and c2 == 'Total' and c3 is not None:
        current_market = str(c1)
        markets[current_market] = {'adr': float(c3), 'rev': float(c4), 'revpar': float(c5), 'rns': int(c6), 'occ': float(c7)*100}
        hotels[current_market] = []
    elif current_market and c2 is not None and c3 is not None:
        hotels.setdefault(current_market, []).append({'name': str(c2), 'adr': float(c3), 'rev': float(c4), 'revpar': float(c5), 'rns': int(c6), 'occ': float(c7)*100})

hh_all = []
for mkt, hlist in hotels.items():
    for h in hlist:
        if h['name'].startswith('HH '):
            hh_all.append({'market': mkt, **h})

lines = []

# Find our hotel
our = None
for h in hh_all:
    if 'Hotel-A' in h['name'] and 'New' not in h['name'] and 'Yinshan' not in h['name'] and 'Wuzhong' not in h['name']:
        our = h
        break

# Filter to core properties
core_hh = [h for h in hh_all if h['occ'] > 30 and h['rns'] > 5000]

# ============================================================
# 1. RevPAR Efficiency Quadrants
# ============================================================
lines.append("=" * 80)
lines.append("  RevPAR æ•ˆçŽ‡å››è±¡é™åˆ†æž?)
lines.append("=" * 80)
lines.append("  (æ¨ªè½´: ADR | çºµè½´: Occ | é¢ç§¯: RevPAR)")
lines.append("")
lines.append("  ã€ç¬¬ä¸€è±¡é™: é«˜ADR+é«˜Occ â€?çŽ‹è€…åŒºé—´ã€?)
lines.append(f"  {'é…’åº—':<38} {'ADR':>8} {'Occ':>8} {'RevPAR':>8} {'RNs':>8}")
lines.append("  " + "-" * 72)
q1 = sorted([h for h in core_hh if h['adr'] >= 800 and h['occ'] >= 75], key=lambda h: h['revpar'], reverse=True)
for h in q1[:10]:
    mkt = h['market'].split('/')[-1]
    lines.append(f"  {h['name']:<38} {h['adr']:>8.0f} {h['occ']:>7.1f}% {h['revpar']:>8.0f} {h['rns']:>8d}")

lines.append("")
lines.append("  ã€ç¬¬äºŒè±¡é™? ä¸­ADR+é«˜Occ â€?æ•ˆçŽ‡æ ‡å…µï¼ˆæˆ‘ä»¬è¯¥å­¦è°ï¼‰ã€?)
lines.append(f"  {'é…’åº—':<38} {'ADR':>8} {'Occ':>8} {'RevPAR':>8} {'RNs':>8}")
lines.append("  " + "-" * 72)
q2 = sorted([h for h in core_hh if 550 <= h['adr'] < 800 and h['occ'] >= 75], key=lambda h: h['revpar'], reverse=True)
for h in q2[:10]:
    mkt = h['market'].split('/')[-1]
    flag = " <<<" if h == our else ""
    lines.append(f"  {h['name']:<38} {h['adr']:>8.0f} {h['occ']:>7.1f}% {h['revpar']:>8.0f} {h['rns']:>8d}{flag}")

lines.append("")
lines.append("  ã€æˆ‘ä»¬çš„ä½ç½®ã€?)
lines.append(f"  Hotel-A           ADR: {our['adr']:.0f} | Occ: {our['occ']:.1f}% | RevPAR: {our['revpar']:.0f}")
lines.append(f"  å±žäºŽ: ç¬¬ä¸‰è±¡é™ï¼ˆä¸­ADR+ä¸­Occï¼?)

# ============================================================
# 2. Top Occ hotels in our ADR bracket
# ============================================================
lines.append("")
lines.append("=" * 80)
lines.append("  ADR Â¥600-700 åŒºé—´æ•ˆçŽ‡æŽ’åï¼ˆæˆ‘ä»¬åŒåŒºé—´å¯¹æ‰‹ï¼?)
lines.append("=" * 80)
lines.append(f"{'æŽ’å':>3} {'é…’åº—':<38} {'å¸‚åœº':<14} {'ADR':>8} {'RevPAR':>8} {'Occ':>7} {'æ•ˆçŽ‡':>8} {'RNs':>8}")
lines.append("-" * 88)
same_adr = sorted([h for h in core_hh if 600 <= h['adr'] <= 700], key=lambda h: h['revpar'], reverse=True)
for i, h in enumerate(same_adr, 1):
    eff = h['revpar'] / h['adr']
    mkt = h['market'].split('/')[-1]
    flag = " <<<" if h == our else ""
    lines.append(f"{i:>3} {h['name']:<38} {mkt:<14} {h['adr']:>8.0f} {h['revpar']:>8.0f} {h['occ']:>6.1f}% {eff:>7.2f} {h['rns']:>8d}{flag}")
    
# Rank of Hotel-A
our_rank = next(i for i, h in enumerate(same_adr, 1) if h == our)
lines.append(f"\n  è‹å·žå¸Œå°”é¡¿åœ¨è¯¥åŒºé—´æŽ’å? {our_rank}/{len(same_adr)}")

# ============================================================
# 3. RevPAR gap decomposition
# ============================================================
lines.append("")
lines.append("=" * 80)
lines.append("  RevPAR å·®è·åˆ†è§£ï¼ˆÂ?00-700 ADRåŒºé—´ï¼?)
lines.append("=" * 80)

# Peer average (same ADR bracket, good hotels)
peers = [h for h in hh_all if 600 <= h['adr'] <= 700 and h['occ'] > 60 and h['rns'] > 20000]
peer_avg_occ = sum(h['occ'] for h in peers) / len(peers)
peer_avg_adr = sum(h['adr'] for h in peers) / len(peers)
peer_avg_revpar = sum(h['revpar'] for h in peers) / len(peers)

# Top 3 performers in bracket
top3 = sorted(peers, key=lambda h: h['revpar'], reverse=True)[:3]
top3_avg_occ = sum(h['occ'] for h in top3) / 3
top3_avg_adr = sum(h['adr'] for h in top3) / 3
top3_avg_revpar = sum(h['revpar'] for h in top3) / 3

# Bottom 30% in bracket
peers_sorted_revpar = sorted(peers, key=lambda h: h['revpar'])
bottom3 = peers_sorted_revpar[:3]

my_occ_pct = our['occ']
lines.append(f"\n{'æŒ‡æ ‡':<24} {'è‹å·žå¸Œå°”é¡?:>12} {'åŒºé—´å‡å€?:>10} {'TOP3å‡å€?:>10}")
lines.append("-" * 56)
lines.append(f"{'ADR':<24} {our['adr']:>12.0f} {peer_avg_adr:>10.0f} {top3_avg_adr:>10.0f}")
lines.append(f"{'Occ (%)':<24} {my_occ_pct:>12.1f} {peer_avg_occ:>10.1f} {top3_avg_occ:>10.1f}")
lines.append(f"{'RevPAR':<24} {our['revpar']:>12.0f} {peer_avg_revpar:>10.0f} {top3_avg_revpar:>10.0f}")

occ_gap = top3_avg_occ - my_occ_pct
lines.append(f"\n--- Occå·®è· ({top3_avg_occ:.1f}% - {my_occ_pct:.1f}% = {occ_gap:+.1f}pp) ---")
lines.append(f"æ‹‰å¹³Occ = å¤šå– {our['rns']/my_occ_pct*100*occ_gap/100:,.0f} é—´å¤œ")
lines.append(f"         = å¢žæ”¶ Â¥{our['adr'] * our['rns']/my_occ_pct*100*occ_gap/100 / 1e4:,.0f} ä¸‡ï¼ˆYTDåŠå¹´ï¼?)

lines.append(f"\nADRå·®è· ({top3_avg_adr:.0f} - {our['adr']:.0f} = {top3_avg_adr-our['adr']:+.0f}):")
lines.append(f"æ‹‰å¹³ADR = å¢žæ”¶ Â¥{(top3_avg_adr-our['adr']) * our['rns'] / 1e4:,.0f} ä¸‡ï¼ˆä»…YTDï¼?)

lines.append(f"\nåŒç®¡é½ä¸‹ï¼ˆADRæ‹‰å¹³TOP3 + Occæ‹‰å¹³TOP3ï¼? å¢žæ”¶ Â¥{(top3_avg_adr-our['adr']) * our['rns']/my_occ_pct*100*occ_gap/100 + top3_avg_adr*our['rns']/my_occ_pct*100*occ_gap/100 / 1e4:,.0f} ä¸?)

# Bottom 3 in same bracket for comparison
lines.append("")
lines.append("--- åŒºé—´åº•éƒ¨è¡¨çŽ° ---")
for h in bottom3:
    mkt = h['market'].split('/')[-1]
    lines.append(f"  {h['name']:<38} {h['adr']:>8.0f} {h['occ']:>6.1f}% RevPAR {h['revpar']:.0f} RNs {h['rns']}")

# ============================================================
# 4. The "Golden Ratio" analysis
# ============================================================
lines.append("")
lines.append("=" * 80)
lines.append("  ðŸ† æœ€ä¼˜æ•ˆçŽ‡æ¨¡åž‹ï¼šè°æ˜¯RevPARçŽ‹è€…ï¼Ÿ")
lines.append("=" * 80)
lines.append("RevPAR = ADR Ã— Occï¼Œä¸¤è€…ä¹˜ç§¯æœ€å¤§åŒ–çš„é…’åº—ï¼š")
lines.append("")
lines.append(f"{'æŽ’å':>3} {'é…’åº—':<38} {'ADR':>8} {'Occ':>8} {'RevPAR':>8} {'ADRÃ—Occ':>12}")
lines.append("-" * 76)

by_product = sorted(core_hh, key=lambda h: h['adr'] * h['occ']/100, reverse=True)
for i, h in enumerate(by_product[:15], 1):
    product = h['adr'] * h['occ']/100
    mkt = h['market'].split('/')[-1]
    flag = " <<<" if h == our else ""
    lines.append(f"{i:>3} {h['name']:<38} {h['adr']:>8.0f} {h['occ']:>6.1f}% {h['revpar']:>8.0f} {product:>10.0f}{flag}")

our_product = our['adr'] * our['occ']/100
our_rank_by_product = next(i for i, h in enumerate(by_product, 1) if h == our)
lines.append(f"\n  è‹å·žå¸Œå°”é¡¿ï¼šADRÃ—Occ = {our_product:.0f}ï¼ŒæŽ’å?{our_rank_by_product}/{len(by_product)}")

# ============================================================
# 5. Key takeaway
# ============================================================
lines.append("")
lines.append("=" * 80)
lines.append("  æ ¸å¿ƒç»“è®º")
lines.append("=" * 80)
lines.append(f"""
1. æˆ‘ä»¬åœ¨Â?00-700 ADRåŒºé—´æŽ’ç¬¬{our_rank}/{len(same_adr)}ï¼ˆæŒ‰RevPARï¼?2. åŒç­‰ADRå¯¹æ‰‹çš„å¹³å‡Occ {peer_avg_occ:.1f}%ï¼Œæˆ‘ä»¬{my_occ_pct:.1f}% â€?å·®{peer_avg_occ-my_occ_pct:.1f}pp
3. ADRæˆ‘ä»¬Â¥{our['adr']:.0f} vs åŒºé—´å‡å€¼Â¥{peer_avg_adr:.0f} â€?å…¶å®žè¿˜é«˜äºŽå‡å€¼ï¼
4. æ‰€ä»¥æ ¸å¿ƒé—®é¢˜æ˜¯Occï¼Œä¸æ˜¯ADR

æœ€ç›´æŽ¥çš„æ”¹å–„è·¯å¾„ï¼š
  1ï¸âƒ£ çŸ­æœŸï¼šæ‹‰Occåˆ°åŒºé—´å‡å€¼{peer_avg_occ:.0f}% â†?RevPARÂ¥{our['adr']*peer_avg_occ/100:.0f}ï¼?Â¥{our['adr']*peer_avg_occ/100-our['revpar']:.0f}ï¼?  2ï¸âƒ£ ä¸­æœŸï¼šæ‹‰Occåˆ°åŒºé—´TOP3æ°´å¹³{top3_avg_occ:.0f}% â†?ï¿¥{our['adr']*top3_avg_occ/100:.0f}
  3ï¸âƒ£ é•¿æœŸï¼šADR+OccåŒç®¡é½ä¸‹ â†?ï¿¥{top3_avg_adr*top3_avg_occ/100:.0f}ï¼ˆTOP3å¹³å‡RevPARï¼?""")

outpath = r'C:\Users\Y\.openclaw\workspace\knowledge_center\revpar_deep_analysis.md'
with open(outpath, 'w', encoding='utf-8') as f:
    f.write('\n'.join(lines))
with open(outpath, 'r', encoding='utf-8') as f:
    print(f.read())
