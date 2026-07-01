#!/usr/bin/env python3
"""Analyze GCM YTD China market data - write to file."""
import openpyxl, sys

path = r'C:\Users\Y\.openclaw\workspace\media\incoming\GCM_YTD.xlsx'
wb = openpyxl.load_workbook(path, data_only=True)
ws = wb['Export']

markets = []
hotels = {}
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
    vals = [v for v in row if v is not None]
    if len(vals) < 7:
        continue
    name = str(vals[0])
    hname = str(vals[1]).strip()
    if hname == 'Total':
        markets.append({
            'name': name, 'adr': float(vals[2]), 'rev': float(vals[3]),
            'revpar': float(vals[4]), 'rns': int(vals[5]), 'occ': float(vals[6])*100,
        })
    elif hname != 'Total':
        hotels[name] = hotels.get(name, []) + [{
            'name': hname, 'adr': float(vals[2]), 'rev': float(vals[3]),
            'revpar': float(vals[4]), 'rns': int(vals[5]), 'occ': float(vals[6])*100,
        }]

markets.sort(key=lambda m: m['rev'], reverse=True)

lines = []
lines.append("=" * 80)
lines.append("  GCM YTD å…¨å›½å¸‚åœºæŽ’å (2026 Jan-May)")
lines.append("=" * 80)
lines.append(f"{'æŽ’å':>3} {'å¸‚åœº':<22} {'ADR':>8} {'Rev(ä¸?':>10} {'RevPAR':>8} {'RNs':>8} {'Occ':>7} {'ç»´åº¦':>6}")
lines.append("-" * 68)
for i, m in enumerate(markets, 1):
    rev_m = m['rev'] / 1e4
    dim = f"{rev_m/10:.0f}é—? if rev_m < 5000 else f"{rev_m/50:.0f}äº? if rev_m > 50000 else "ä¸­åž‹"
    flag = " <<<" if "City-X" in m['name'] else ""
    lines.append(f"{i:>3} {m['name']:<22} {m['adr']:>8.0f} {rev_m:>10.1f} {m['revpar']:>8.0f} {m['rns']:>8d} {m['occ']:>6.1f}%{flag}")

lines.append("")
lines.append("=" * 80)
lines.append("  è‹å·žå¸‚åœºé…’åº—æ˜Žç»†")
lines.append("=" * 80)
lines.append(f"{'å“ç‰Œ/é…’åº—':<42} {'ADR':>8} {'Rev(ä¸?':>10} {'RevPAR':>8} {'RNs':>7} {'Occ':>7} {'ä»½é¢/ADRæ¯?:>10}")
lines.append("-" * 86)

# Find City-X section and collect hotels by reading cell directly
sz_hotels = []
current_market = None
for r in range(2, ws.max_row + 1):
    col1 = ws.cell(r, 1).value
    col2 = ws.cell(r, 2).value
    col3 = ws.cell(r, 3).value
    col4 = ws.cell(r, 4).value
    col5 = ws.cell(r, 5).value
    col6 = ws.cell(r, 6).value
    col7 = ws.cell(r, 7).value
    
    if col1 is not None and col2 == 'Total' and col3 is not None:
        current_market = str(col1)
        if 'City-X' in current_market:
            sz_hotels.append(('ã€å¸‚åœºæ€»è®¡ã€?, float(col3), float(col4), float(col5), int(col6), float(col7)*100))
            sz_total_rev = float(col4)
        continue
    if current_market and 'City-X' in current_market and col2 is not None and col3 is not None:
        adr = float(col3); rev = float(col4); revpar = float(col5)
        rns = int(col6); occ = float(col7) * 100
        sz_hotels.append((str(col2), adr, rev, revpar, rns, occ))

for name, adr, rev, revpar, rns, occ in sz_hotels:
    share = rev / sz_total_rev * 100
    adr_ratio = adr / (sz_hotels[0][1])  # ratio vs market avg ADR
    flag = " <<< è‹å·žå¸Œå°”é¡? if 'Hotel-A' in name and 'New' not in name and 'Yinshan' not in name and 'Wuzhong' not in name else ""
    lines.append(f"{name:<42} {adr:>8.0f} {rev/1e4:>10.1f} {revpar:>8.0f} {rns:>7d} {occ:>6.1f}% {share:>6.1f}%/x{adr_ratio:.2f}{flag}")

outpath = r'C:\Users\Y\.openclaw\workspace\knowledge_center\gcm_ytd_report.md'
with open(outpath, 'w', encoding='utf-8') as f:
    f.write('\n'.join(lines))
print(f"Report saved: {outpath}")
print('\n'.join(lines))
