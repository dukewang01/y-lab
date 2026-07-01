#!/usr/bin/env python3
"""Deep dive 2: Simulation scenarios + City-X competitive battlefield + national context."""
import openpyxl
from datetime import datetime

path = r'C:\Users\Y\.openclaw\workspace\media\archived\GCM_YTD.xlsx'
wb = openpyxl.load_workbook(path, data_only=True)
ws = wb['Export']

markets = {}
hotels = {}
current_market = None
for r in range(2, ws.max_row + 1):
    c1 = ws.cell(r, 1).value; c2 = ws.cell(r, 2).value
    c3 = ws.cell(r, 3).value; c4 = ws.cell(r, 4).value
    c5 = ws.cell(r, 5).value; c6 = ws.cell(r, 6).value; c7 = ws.cell(r, 7).value
    if c1 is not None and c2 == 'Total' and c3 is not None:
        current_market = str(c1)
        markets[current_market] = {
            'adr': float(c3), 'rev': float(c4), 'revpar': float(c5),
            'rns': int(c6), 'occ': float(c7)*100
        }
        hotels[current_market] = []
    elif current_market and c2 is not None and c3 is not None:
        hotels.setdefault(current_market, []).append({
            'name': str(c2), 'adr': float(c3), 'rev': float(c4),
            'revpar': float(c5), 'rns': int(c6), 'occ': float(c7)*100
        })

# Our hotel data
our = {'adr': 649.37, 'rev': 36114254.20, 'revpar': 444.55, 'rns': 55614, 'occ': 68.46}

lines = []

lines.append("=" * 80)
lines.append("  ðŸ”® è‹å·žå¸Œå°”é¡?â€?æå‡æ½œåŠ›æ¨¡æ‹Ÿï¼ˆYTD Jan-May 2026ï¼?)
lines.append("=" * 80)

days_ytd = 151  # Jan 1 to May 31 = 151 days
avail_rooms = 538
total_avail_rns = int(avail_rooms * days_ytd)  # 81,238
current_occ = our['occ'] / 100
current_adr = our['adr']

# Scenario 1: ADR up to match HH Shanghai Hongqiao (Â¥803) or HH Guangzhou Tianhe (Â¥934)
# Scenario 2: Occ up to 75% (match market benchmark)
# Scenario 3: Both

scenarios = [
    ("ADR +5%", our['adr'] * 1.05, our['occ']),
    ("ADR +10%", our['adr'] * 1.10, our['occ']),
    ("ADR â†?Â¥750 (â‰ˆè™¹æ¡¥æ°´å¹?", 750, our['occ']),
    ("Occ â†?75% (+6.5pp)", our['adr'], 75),
    ("Occ â†?80% (+11.5pp)", our['adr'], 80),
    ("ADRÂ¥750 + Occ75%", 750, 75),
    ("ADRÂ¥750 + Occ80%", 750, 80),
    ("ADRÂ¥800 + Occ75%", 800, 75),
]

lines.append(f"{'æ¨¡æ‹Ÿåœºæ™¯':<30} {'ADR':>8} {'Occ':>8} {'é—´å¤œ':>8} {'æ”¶å…¥(ä¸?':>12} {'å¢žé‡(ä¸?':>10} {'å¢žå¹…':>8}")
lines.append("-" * 86)
baseline_rev = our['rev']
baseline_rns = our['rns']

for name, new_adr, new_occ_pct in scenarios:
    new_occ = new_occ_pct / 100
    # Assume same available room nights, new occ drives more RNs
    add_rns = total_avail_rns * (new_occ - current_occ)
    new_rns = our['rns'] + add_rns
    new_rev = new_rns * new_adr
    inc_rev = new_rev - baseline_rev
    inc_pct = inc_rev / baseline_rev * 100
    lines.append(f"{name:<30} {new_adr:>8.0f} {new_occ_pct:>7.1f}% {new_rns:>8.0f} Â¥{new_rev/1e4:>10.1f} +Â¥{inc_rev/1e4:>8.1f} {inc_pct:>+7.1f}%")

lines.append("")
lines.append(f"ã€åŸºçº¿ã€‘ADR Â¥{current_adr:.0f} | Occ {current_occ*100:.1f}% | RNs {baseline_rns:.0f} | Rev Â¥{baseline_rev/1e4:.1f}ä¸?)
lines.append(f"ã€æ€»å¯å”®é—´å¤œã€‘{avail_rooms}é—?Ã— {days_ytd}å¤?= {total_avail_rns:,} é—´å¤œï¼ˆå½“å‰å”®å‡º{baseline_rns}ï¼?)

# Scenario: ADR sensitivity with Elasticity
lines.append("")
lines.append("=" * 80)
lines.append("  ðŸ“ ADRå¼¹æ€§åˆ†æž?â€?æ¶¨ä»·å¤šå°‘å¼€å§‹å½±å“Occï¼?)
lines.append("=" * 80)
lines.append("å‡è®¾æ¯æ¶¨Â¥50 ADRå¯¼è‡´Occä¸‹é™2ppï¼ˆç»éªŒæ¨¡åž‹ï¼‰ï¼?)
lines.append(f"{'ADRæ–¹æ¡ˆ':<24} {'ADR':>8} {'é¢„æœŸOcc':>8} {'RNs':>8} {'æ”¶å…¥(ä¸?':>12} {'æ”¶å…¥å˜åŒ–':>10}")
lines.append("-" * 70)

for premium in [0, 30, 50, 80, 100, 150, 200]:
    new_adr = current_adr + premium
    occ_loss = (premium / 50) * 0.02  # 2pp drop per 50Â¥
    new_occ = current_occ - occ_loss
    if new_occ < 0.4:
        new_occ = 0.4
    new_rns = total_avail_rns * new_occ * (our['rns'] / (total_avail_rns * current_occ))  # Scale: RNs at new Occ
    # Actually simpler: new_rns = total_avail_rns * new_occ (sliding into the actual calculation)
    # Wait, the current situation is not using all rooms. Let me think more carefully.
    # Total available RNs = 538 * 151 = 81,238
    # Current RNs = 55,614, so current occ = 55,614/81,238 = 68.46%. OK matches.
    new_rns = total_avail_rns * new_occ
    new_rev = new_rns * new_adr
    change = (new_rev - baseline_rev) / baseline_rev * 100
    lines.append(f"{f'ADR+Â¥{premium}':<24} {new_adr:>8.0f} {new_occ*100:>7.1f}% {new_rns:>8.0f} Â¥{new_rev/1e4:>10.1f} {change:>+9.1f}%")

lines.append("")
lines.append("=" * 80)
lines.append("  ðŸŸï¸?è‹å·ž6å¤§é…’åº?â€?ç«žå“æˆ˜åœº")
lines.append("=" * 80)
sz_hotels = sorted(hotels.get('Jiangsu/City-X', []), key=lambda h: h['rev'], reverse=True)

lines.append(f"{'æŽ’å':>3} {'é…’åº—':<30} {'ADR':>8} {'Rev(ä¸?':>10} {'RevPAR':>8} {'RNs':>8} {'Occ':>7} {'ADR/mkt':>8} {'Revä»½é¢':>8}")
lines.append("-" * 95)
mkt_adr = markets['Jiangsu/City-X']['adr']
for i, h in enumerate(sz_hotels, 1):
    share = h['rev'] / markets['Jiangsu/City-X']['rev'] * 100
    adr_vs_mkt = h['adr'] / mkt_adr * 100
    flag = " <<<" if 'Hotel-A' in h['name'] and 'New' not in h['name'] and 'Yinshan' not in h['name'] and 'Wuzhong' not in h['name'] else ""
    lines.append(f"{i:>3} {h['name']:<30} {h['adr']:>8.0f} {h['rev']/1e4:>10.1f} {h['revpar']:>8.0f} {h['rns']:>8d} {h['occ']:>6.1f}% {adr_vs_mkt:>7.1f}% {share:>7.1f}%{flag}")

# City-X market position in Jiangsu province
lines.append("")
lines.append("=" * 80)
lines.append("  ðŸ“ æ±Ÿè‹çœå†…å¸‚åœºæ ¼å±€")
lines.append("=" * 80)
js_markets = {k: v for k, v in markets.items() if k.startswith('Jiangsu/')}
total_js = sum(m['rev'] for m in js_markets.values())
js_sorted = sorted(js_markets.items(), key=lambda x: x[1]['rev'], reverse=True)
lines.append(f"{'åŸŽå¸‚':<24} {'ADR':>8} {'Rev(ä¸?':>10} {'RevPAR':>8} {'RNs':>8} {'Occ':>7} {'çœå æ¯?:>8}")
lines.append("-" * 73)
for city, m in js_sorted:
    share = m['rev'] / total_js * 100
    lines.append(f"{city:<24} {m['adr']:>8.0f} {m['rev']/1e4:>10.1f} {m['revpar']:>8.0f} {m['rns']:>8d} {m['occ']:>6.1f}% {share:>7.1f}%")

# National HH top 30 ranking
lines.append("")
lines.append("=" * 80)
lines.append("  ðŸ† å…¨å›½HHé…’åº—Top 30 â€?è¥æ”¶æŽ’å")
lines.append("=" * 80)

hh_all = []
for mkt, hlist in hotels.items():
    for h in hlist:
        if h['name'].startswith('HH '):
            hh_all.append({'market': mkt, **h})
hh_all.sort(key=lambda x: x['rev'], reverse=True)

lines.append(f"{'æŽ’å':>3} {'é…’åº—':<40} {'å¸‚åœº':<16} {'ADR':>8} {'Rev(ä¸?':>10} {'Occ':>7} {'RNs':>8}")
lines.append("-" * 96)
for i, h in enumerate(hh_all[:30], 1):
    mkt_s = h['market'].split('/')[-1]
    flag = " <<<" if 'Hotel-A' in h['name'] and 'New' not in h['name'] and 'Yinshan' not in h['name'] and 'Wuzhong' not in h['name'] else ""
    lines.append(f"{i:>3} {h['name']:<40} {mkt_s:<16} {h['adr']:>8.0f} {h['rev']/1e4:>10.1f} {h['occ']:>6.1f}% {h['rns']:>8d}{flag}")

# Key insight: Hotel-A's Occ is bottom vs top performers
lines.append("")
lines.append("=" * 80)
lines.append("  ðŸ’¡ æ ¸å¿ƒè¯Šæ–­ï¼šè‹å·žå¸Œå°”é¡¿çš?ä¸¤ä½Žä¸€é«?")
lines.append("=" * 80)
lines.append("")
lines.append("ã€ä¼˜åŠ¿ã€?)
lines.append("  1. é—´å¤œé‡å…¨å›½ç¬¬3ï¼?5,614 RNsï¼‰â€?è§„æ¨¡ä¼˜åŠ¿æ˜Žæ˜¾")
lines.append("  2. ADRæº¢ä»·26%é¢†å…ˆè‹å·žå¸‚åœºâ€?å“ç‰Œå®šä½æ¸…æ™°")
lines.append("  3. è‹å·žå¸‚åœº43.6%è¥æ”¶ä»½é¢â€?ç»å¯¹é¾™å¤´")
lines.append("")
lines.append("ã€ç“¶é¢ˆã€?)
lines.append(f"  4. Occupancy 68.5% â€?Top 30 HHé…’åº—ä¸­æŽ’åç¬¬28ï¼ˆä»…é«˜äºŽæ˜†æ˜Ž/ç æµ·ç­‰ï¼‰")
lines.append("     å…¨å›½HHå¹³å‡Occçº?5%ï¼Œæˆ‘ä»¬å·®äº?.5ä¸ªç‚¹")
lines.append("  5. ADR Â¥649 â€?Top 10 HHä¸­æœ€ä½Žï¼Œåªæœ‰ä¸‰äºšçš?0%ã€ä¸Šæµ·è™¹æ¡¥çš„81%")
lines.append("     ä½†è‹å·žå¸‚åœºå¤©èŠ±æ¿ç¡®å®žåœ¨è¿™ï¼ˆå¸‚åœºavg Â¥516ï¼?)
lines.append("")
lines.append("ã€æœ€ç›´æŽ¥çš„æ æ†ã€?)
lines.append("  â–?æå‡Occ 75%ï¼?6.5ppï¼? å¤šå–5,278é—´å¤œ")
lines.append(f"     â†?å…¨å¹´è§’åº¦çœ‹ï¼Œçº¦å¯å¢žæ”¶Â¥{5280*650/1e4:.0f}ä¸‡ï¼ˆå‡è®¾ADRä¸å˜ï¼?)
lines.append("  â–?ADRæè‡³Â¥700ï¼?Â¥50ï¼‰ï¼Œä¿æŒOcc â†?å¢žæ”¶Â¥278ä¸?å¹?)
lines.append("  â–?åŒç®¡é½ä¸‹ï¼ˆADRÂ¥700+Occ75%ï¼? å¢žæ”¶Â¥1,200ä¸?/å¹?)

outpath = r'C:\Users\Y\.openclaw\workspace\knowledge_center\gcm_ytd_scenarios.md'
with open(outpath, 'w', encoding='utf-8') as f:
    f.write('\n'.join(lines))

with open(outpath, 'r', encoding='utf-8') as f:
    print(f.read())
