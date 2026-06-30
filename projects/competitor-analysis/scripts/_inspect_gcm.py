#!/usr/bin/env python3
"""Inspect GCM YTD Excel file."""
import openpyxl

path = r'C:\Users\Duke Wang\.openclaw\workspace\media\incoming\GCM_YTD.xlsx'
wb = openpyxl.load_workbook(path, data_only=True)
print('Sheets:', wb.sheetnames)
for name in wb.sheetnames:
    ws = wb[name]
    print(f'\n=== Sheet: {name} ({ws.max_row} rows x {ws.max_column} cols) ===')
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=min(15, ws.max_row), values_only=True)):
        vals = [v for v in row if v is not None]
        if vals and any(v is not None and ('su' in str(v).lower() or '总' in str(v) or 'Total' in str(v)) for v in vals):
            print(f'  R{i+1}: {vals[:15]}')
