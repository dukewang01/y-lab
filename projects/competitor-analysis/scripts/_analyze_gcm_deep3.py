#!/usr/bin/env python3
"""GCM Deep Dive Round 3: Brand power, efficiency patterns, market structure."""
import openpyxl

path = r'C:\Users\Y\.openclaw\workspace\media\archived\GCM_YTD.xlsx'
wb = openpyxl.load_workbook(path, data_only=True)
ws = wb['Export']

markets = {}
hotels = {}
current_market = None
for r in range(2, ws.max_row + 1):
    c1 = ws.cell(r, 1).value; c2 = ws.cell(r, 2).value
    c3 = ws.cell(r, 3).value; c4 = ws.cell(r, 4).value
    c5 = ws.cell(r, 5).value; c6 = ws.cell(r, 6).value; c7 = ws.cell(r, 7).value
    if c1 is not None and c2 == 'Total' and c3 is not None:
        current_market = str(c1)
        markets[current_market] = {
            'adr': float(c3), 'rev': float(c4), 'revpar': float(c5),
            'rns': int(c6), 'occ': float(c7)*100
        }
        hotels[current_market] = []
    elif current_market and c2 is not None and c3 is not None:
        hotels.setdefault(current_market, []).append({
            'name': str(c2), 'adr': float(c3), 'rev': float(c4),
            'revpar': float(c5), 'rns': int(c6), 'occ': float(c7)*100
        })

lines = []

# ===================================================================
# 1. ADR vs RevPAR scatter: Find the "efficiency champions"
# ===================================================================
lines.append("=" * 80)
lines.append("  ðŸ“Š æ•ˆçŽ‡å›¾è°±ï¼šè°åœ?RevPAR ä¸Šè·‘èµ¢äº† ADRï¼?)
lines.append("=" * 80)
lines.append("ï¼ˆRevPAR Ã· ADR = Occ â†?è¶ŠæŽ¥è¿?è¶Šå¥½ï¼Œä½†ADRä¹Ÿé‡è¦ï¼‰")
lines.append("")
lines.append(f"{'æŽ’å':>3} {'é…’åº—':<38} {'å¸‚åœº':<18} {'ADR':>8} {'RevPAR':>8} {'Occ':>7} {'æ•ˆçŽ‡':>8}")
lines.append("-" * 90)

hh_all = []
for mkt, hlist in hotels.items():
    for h in hlist:
        if h['name'].startswith('HH '):
            hh_all.append({'market': mkt, **h})

# "Efficiency Ratio" = RevPAR/ADR (essentially Occ) but also consider absolute RevPAR
# Let's find hotels with BOTH high ADR and high RevPAR
for h in hh_all:
    h['efficiency'] = h['revpar'] / h['adr']

# Top by absolute RevPAR
revpar_sorted = sorted(hh_all, key=lambda x: x['revpar'], reverse=True)
for i, h in enumerate(revpar_sorted[:15], 1):
    flag = " <<<" if 'Hotel-A' in h['name'] and 'New' not in h['name'] and 'Yinshan' not in h['name'] and 'Wuzhong' not in h['name'] else ""
    mkt_s = h['market'].split('/')[-1]
    lines.append(f"{i:>3} {h['name']:<38} {mkt_s:<18} {h['adr']:>8.0f} {h['revpar']:>7.0f} {h['occ']:>6.1f}% {h['efficiency']:>7.2f}{flag}")

# Efficiency bottom (low RevPAR given ADR)
lines.append("")
lines.append("  â†?Bottom 10 by RevPAR efficiency...")
bottom_eff = sorted(hh_all, key=lambda x: x['revpar'])
for i, h in enumerate(bottom_eff[:10], 1):
    mkt_s = h['market'].split('/')[-1]
    lines.append(f"{i:>3} {h['name']:<38} {mkt_s:<18} {h['adr']:>8.0f} {h['revpar']:>7.0f} {h['occ']:>6.1f}% {h['efficiency']:>7.2f}")

# ===================================================================
# 2. Brand premium: HH vs GI vs DT in each market
# ===================================================================
lines.append("")
lines.append("=" * 80)
lines.append("  ðŸ·ï¸?å“ç‰Œæº¢ä»·åŠ›ï¼šHH vs GI vs DTï¼ˆåŒåŸŽç«žå“ï¼‰")
lines.append("=" * 80)
lines.append(f"{'å¸‚åœº':<20} {'HH ADR':>8} {'GI ADR':>8} {'DT ADR':>8} {'HHæº¢ä»·vsGI':>12} {'HHæº¢ä»·vsDT':>12}")
lines.append("-" * 68)

brand_data = {}
for mkt, hlist in hotels.items():
    for h in hlist:
        code = h['name'].split()[0] if ' ' in h['name'] else h['name']
        key = code[:2]
        brand_data.setdefault(mkt, {}).setdefault(key, []).append(h)

for mkt in sorted(brand_data.keys()):
    brands = brand_data[mkt]
    hh = brands.get('HH', [])
    gi = brands.get('GI', [])
    dt = brands.get('DT', [])
    # Exclude non-flagship HH (like Yinshan Lake, New District)
    hh_main = [h for h in hh if len(h['name'].split()) <= 3 or 'Yalong' in h['name'] or 'City Center' in h['name'] or 'Tianhe' in h['name'] or 'Futian' in h['name']]
    if not hh_main:
        hh_main = hh[:1]
    
    hh_adr = sum(h['adr'] for h in hh_main) / len(hh_main) if hh_main else 0
    gi_adr = sum(h['adr'] for h in gi) / len(gi) if gi else 0
    dt_adr = sum(h['adr'] for h in dt) / len(dt) if dt else 0
    
    pr_vs_gi = (hh_adr / gi_adr - 1) * 100 if gi_adr else None
    pr_vs_dt = (hh_adr / dt_adr - 1) * 100 if dt_adr else None
    
    hi_mkt = " <<< City-X" if 'City-X' in mkt else ""
    if gi_adr or dt_adr:
        gi_s = f"{gi_adr:>8.0f}" if gi_adr else "   N/A"
        dt_s = f"{dt_adr:>8.0f}" if dt_adr else "   N/A"
        pgi = f"{pr_vs_gi:>+10.0f}%" if pr_vs_gi is not None else "      N/A"
        pdt = f"{pr_vs_dt:>+10.0f}%" if pr_vs_dt is not None else "      N/A"
        mkt_short = mkt.split('/')[-1] if '/' in mkt else mkt
        lines.append(f"{mkt_short:<20} {hh_adr:>8.0f} {gi_s:>8} {dt_s:>8} {pgi}{pdt}{hi_mkt}")

# ===================================================================
# 3. Market density analysis: does more competition hurt ADR?
# ===================================================================
lines.append("")
lines.append("=" * 80)
lines.append("  ðŸ™ï¸?å¸‚åœºç»“æž„ï¼šé…’åº—å¯†åº?vs ä»·æ ¼è¡¨çŽ°")
lines.append("=" * 80)
lines.append(f"{'å¸‚åœº':<22} {'é…’åº—æ•?:>6} {'å¸‚åœºADR':>10} {'æ——èˆ°ADR':>10} {'æ——èˆ°ä»½é¢':>10} {'Occ':>7}")
lines.append("-" * 65)

density = []
for mkt, hlist in hotels.items():
    n = len(hlist)
    hh_main = [h for h in hlist if h['name'].startswith('HH ')]
    gi_main = [h for h in hlist if h['name'].startswith('GI ')]
    dt_main = [h for h in hlist if h['name'].startswith('DT ')]
    
    # Primary HH (first listed)
    primary = hh_main[0]['adr'] if hh_main else 0
    primary_share = hh_main[0]['rev'] / markets[mkt]['rev'] * 100 if hh_main and markets[mkt]['rev'] else 0
    
    density.append({
        'market': mkt,
        'hotels': n,
        'hh_count': len(hh_main),
        'gi_count': len(gi_main),
        'dt_count': len(dt_main),
        'market_adr': markets[mkt]['adr'],
        'primary_adr': primary,
        'primary_share': primary_share,
        'occ': markets[mkt]['occ']
    })

density.sort(key=lambda x: x['hotels'], reverse=True)

for d in density[:20]:
    flag = " <<<" if 'City-X' in d['market'] else ""
    mkt_s = d['market'].split('/')[-1] if '/' in d['market'] else d['market']
    lines.append(f"{mkt_s:<22} {d['hotels']:>6d} {d['market_adr']:>10.0f} {d['primary_adr']:>10.0f} {d['primary_share']:>9.1f}% {d['occ']:>6.1f}%{flag}")

# ===================================================================
# 4. The "sweet spot" analysis: find optimal ADR zone
# ===================================================================
lines.append("")
lines.append("=" * 80)
lines.append("  ðŸ“ˆ æœ€ä½³åŒºé—´ï¼šå…¨å›½HHé…’åº—çš„ADRåˆ†å¸ƒä¸Žè¥æ”¶æ•ˆçŽ?)
lines.append("=" * 80)

# Group HH hotels by ADR bracket
brackets = [(0, 400), (400, 500), (500, 600), (600, 700), (700, 800), 
            (800, 900), (900, 1000), (1000, 1200), (1200, 1500), (1500, 9999)]
bracket_labels = ['<400', '400-500', '500-600', '600-700', '700-800', 
                  '800-900', '900-1000', '1000-1200', '1200-1500', '>1500']
bracket_data = {l: {'count': 0, 'adr': 0, 'rev': 0, 'occ': 0, 'revpar': 0, 'rns': 0} for l in bracket_labels}

for h in hh_all:
    for i, (lo, hi) in enumerate(brackets):
        if lo <= h['adr'] < hi:
            l = bracket_labels[i]
            bracket_data[l]['count'] += 1
            bracket_data[l]['adr'] += h['adr']
            bracket_data[l]['rev'] += h['rev']
            bracket_data[l]['occ'] += h['occ']
            bracket_data[l]['revpar'] += h['revpar']
            bracket_data[l]['rns'] += h['rns']
            break

lines.append(f"{'ADRåŒºé—´':<14} {'é…’åº—æ•?:>6} {'å‡ADR':>8} {'å‡Rev':>10} {'å‡Occ':>8} {'å‡RevPAR':>10} {'æ€»é—´å¤?:>10}")
lines.append("-" * 66)
for l in bracket_labels:
    d = bracket_data[l]
    if d['count'] > 0:
        avg_adr = d['adr'] / d['count']
        avg_rev = d['rev'] / d['count'] / 1e4
        avg_occ = d['occ'] / d['count']
        avg_revpar = d['revpar'] / d['count']
        lines.append(f"{l:<14} {d['count']:>6d} {avg_adr:>8.0f} Â¥{avg_rev:>8.1f}ä¸?{avg_occ:>7.1f}% {avg_revpar:>9.0f} {d['rns']:>10d}")

# ===================================================================
# 5. Same Hotel Group comparison: Hotel-A vs peers with similar city tier
# ===================================================================
lines.append("")
lines.append("=" * 80)
lines.append("  ðŸŽ¯ åŒçº§åŸŽå¸‚å¯¹æ ‡ï¼šéžä¸€çº¿åŸŽå¸‚HHé…’åº—å¯¹æ¯”")
lines.append("=" * 80)
lines.append(f"{'æŽ’å':>3} {'é…’åº—':<34} {'åŸŽå¸‚':<14} {'ADR':>8} {'RevPAR':>8} {'Occ':>7} {'RNs':>8}")
lines.append("-" * 80)

tier2 = ['Guangzhou','Nanjing','Shenzhen','Hangzhou','Shenzhen','Chengdu','Wuhan',
         'Xi\'an','Zhengzhou','Changsha','Foshan','Shijiazhuang','Haikou','Urumqi']
# Actually Shenzhen/Guangzhou are tier 1.5, let me be more specific - "provincial capitals + single hotel markets"
# Markets with one primary HH and clear tier-2 characteristics
tier2_markets = []

for h in hh_all:
    mkt_full = h['market']
    mkt_city = mkt_full.split('/')[-1] if '/' in mkt_full else mkt_full
    # Skip tier-1 cities explicitly
    if mkt_city in ['Shanghai','Beijing','Sanya']:
        continue
    # Skip Yalong Bay (Sanya)
    if 'Yalong' in h['name'] or 'Yuxi' in h['name']:
        continue
    # Keep primary HH in non-T1 markets
    if not any(x in h['name'] for x in ['Yinshan','New District','Wuzhong','Baiyun', 
                                          'Science City','Futian','Shekou',
                                          'Capital Airport','Liangjiang',
                                          'Riverside','Niushoushan',
                                          'Tongzhou','Zhangjiang','Hongqiao']):
        if 'City Center' in h['name'] or 'Wangfujing' in h['name']:
            continue  # Skip T1 specialty locations
        tier2_markets.append(h)

tier2_markets.sort(key=lambda x: x['revpar'], reverse=True)

for i, h in enumerate(tier2_markets, 1):
    mkt_city = h['market'].split('/')[-1] if '/' in h['market'] else h['market']
    flag = " <<<" if 'Hotel-A' in h['name'] and 'New' not in h['name'] and 'Yinshan' not in h['name'] and 'Wuzhong' not in h['name'] else ""
    lines.append(f"{i:>3} {h['name']:<34} {mkt_city:<14} {h['adr']:>8.0f} {h['revpar']:>8.0f} {h['occ']:>6.1f}% {h['rns']:>8d}{flag}")

# ===================================================================
# 6. Final insight: The Occ gap analysis
# ===================================================================
lines.append("")
lines.append("=" * 80)
lines.append("  ðŸ”Ž æ ¸å¿ƒè¿½é—®ï¼šä¸ºä»€ä¹ˆè‹å·žOccåªæœ‰68.5%ï¼?)
lines.append("=" * 80)
lines.append("")
lines.append("çœ‹æ•°æ®è¯´è¯ï¼š")
lines.append(f"  â–?è‹å·žå¸Œå°”é¡?68.5% vs è‹å·žå¸‚åœºå¹³å‡ 60.1% â€?é«˜äºŽå¸‚åœº8.4ppï¼Œä¸é”?)
lines.append(f"  â–?è‹å·žå¸Œå°”é¡?68.5% vs å…¨å›½HHå¹³å‡ ~75% â€?ä½†ä½ŽäºŽåŒè¡?.5pp")
lines.append(f"  â–?è‹å·žå¸Œå°”é¡?68.5% vs Yinshan Lake 68.9% â€?æœ¬åœ°ç«žå“å±…ç„¶è¿˜é«˜0.4pp")
lines.append("")
lines.append("å¯èƒ½çš„åŽŸå› ï¼š")
lines.append("  1ï¸âƒ£ å‘¨ä¸­/å‘¨æœ«ç»“æž„å¤±è¡¡ â€?å‘¨ä¸€è‡³å‘¨å››å•†åŠ¡éœ€æ±‚ä¸è¶³ï¼Ÿ")
lines.append("  2ï¸âƒ£ ä¼šå±•/å…¬å¸åè®®å®¢æˆ·æ¸—é€çŽ‡ä¸å¤Ÿ â€?è‹å·žå·¥ä¸šå›­åŒºå†…å…¬å¸åè®®ä»·åä¿å®ˆï¼Ÿ")
lines.append("  3ï¸âƒ£ ç«žäº‰åˆ†æµ â€?Yinshan Lake(Â¥510) ä»¥æ›´ä½Žä»·æ ¼æŠ¢äº†éƒ¨åˆ†é‡")
lines.append("  4ï¸âƒ£ å¯èƒ½å›¢é˜Ÿ/æ—…è¡Œç¤¾åˆä½œä¸å¤Ÿæ·±")
lines.append("")
lines.append("è·Ÿå…¨å›½HHå¯¹æ ‡é…’åº—æ¯”ï¼š")
# Find peer: same-ish ADR level, more Occ
peers = sorted([h for h in hh_all if 600 <= h['adr'] <= 750], key=lambda x: x['occ'], reverse=True)
lines.append(f"{'é…’åº—':<40} {'ADR':>8} {'Occ':>8} {'RevPAR':>8} {'RNs':>8}")
lines.append("-" * 72)
for h in peers[:8]:
    lines.append(f"{h['name']:<40} {h['adr']:>8.0f} {h['occ']:>7.1f}% {h['revpar']:>8.0f} {h['rns']:>8d}")

lines.append("")
lines.append("=== ç»“è®º ===")
lines.append("  è‹å·žå¸Œå°”é¡¿çš„æ”¶å…¥çŸ­æ¿åœ¨Occè€ŒéžADRã€‚å‚è€ƒåŒç­‰ADRå¯¹æ‰‹(Â¥600-750)ï¼?)
lines.append("  å®ƒä»¬Occæ™®éåœ?5-85%ï¼Œæˆ‘ä»?8.5%æ˜Žæ˜¾åä½Žã€?)
lines.append("  Occæå‡6.5ppåˆ?5% = å¤šé”€5,278é—´å¤œ = Â¥343ä¸‡å¢žæ”¶ï¼ˆYTDï¼?)
lines.append("  å…¨å¹´çœ‹è¿™å°†è¿‘Â¥1,000ä¸‡çš„å¢žé‡æ½œåŠ›ã€?)

outpath = r'C:\Users\Y\.openclaw\workspace\knowledge_center\gcm_ytd_deep3.md'
with open(outpath, 'w', encoding='utf-8') as f:
    f.write('\n'.join(lines))

with open(outpath, 'r', encoding='utf-8') as f:
    content = f.read()

# Print without emojis for terminal
for line in content.split('\n'):
    # Try to print each line, skip if fails
    try:
        print(line.replace('ðŸ“Š','').replace('ðŸ·ï¸?,'').replace('ðŸ™ï¸?,'').replace('ðŸ“ˆ','').replace('ðŸŽ¯','').replace('ðŸ”Ž','').replace('ðŸ”¥',''))
    except:
        pass
