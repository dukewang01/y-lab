#!/usr/bin/env python3
import openpyxl, json, os, sys
sys.stdout.reconfigure(encoding='utf-8')

BASE = os.path.dirname(os.path.abspath(__file__))
indir = r'C:\Users\Duke Wang\.openclaw\media\inbound'
fn = [f for f in os.listdir(indir) if 'HOE00033' in f or '1a36370d' in f][0]
fp = os.path.join(indir, fn)
wb = openpyxl.load_workbook(fp, data_only=True)
ws = wb.active

items = []
for r in range(4, 44):
    seq = ws.cell(r, 1).value
    if not seq: break
    try:
        qty_str = str(ws.cell(r, 7).value or '0').strip()
        qty = int(float(qty_str)) if qty_str else 0
        if qty == 0: continue
        items.append({
            'seq': int(seq),
            'name': str(ws.cell(r, 2).value or '').strip(),
            'brand': str(ws.cell(r, 3).value or '').strip(),
            'model': str(ws.cell(r, 4).value or '').strip(),
            'maker': str(ws.cell(r, 5).value or '').strip(),
            'unit': str(ws.cell(r, 6).value or '').strip(),
            'qty': qty,
        })
    except: pass

total_qty = sum(item['qty'] for item in items)
brands = {}
for item in items:
    b = item['brand']
    if b not in brands: brands[b] = {'count':0, 'qty':0}
    brands[b]['count'] += 1
    brands[b]['qty'] += item['qty']

print(f'日韩餐具: {len(items)}项, {total_qty}件, 总价¥52,200')
print(f'\n品牌:')
for b, d in sorted(brands.items(), key=lambda x: -x[1]['qty']):
    maker = items[[i['brand'] for i in items].index(b)]['maker'] if b in [i['brand'] for i in items] else ''
    print(f'  {b:6s}: {d["count"]}种 {d["qty"]}件')

# 入库
fb_fp = os.path.join(BASE, "fb_graph.json")
fb = json.load(open(fb_fp, 'r', encoding='utf-8'))
es = fb.get('entities', [])
existing_ids = set(e.get('id','') for e in es)

# 先清除之前仅1项的残次记录
for eid in ['HOE_VENDOR_JAPANESE_001', 'HOE_CONTRACT_JAPANESE_001']:
    es[:] = [e for e in es if e.get('id') != eid]
existing_ids = set(e.get('id','') for e in es)
for i in range(1, 50):
    es[:] = [e for e in es if e.get('id') != f'HOE_ITEM_JAPAN_{i:03d}']
existing_ids = set(e.get('id','') for e in es)

hoes = [
    {"id": "HOE_VENDOR_JAPANESE_001", "type": "hoe_vendor", "label": "日韩式餐具供应商",
     "category": "设备器具", "status": "合作中", "contract_total": "¥52,200",
     "import_date": "2026-05-14"},
    {"id": "HOE_CONTRACT_JAPANESE_001", "type": "hoe_contract", "label": "HOE00033 日韩式餐具合同",
     "vendor_id": "HOE_VENDOR_JAPANESE_001", "contract_type": "供应合同",
     "category": "设备器具", "status": "合作中", "items_count": len(items),
     "total_qty": total_qty, "total_amount": 52200,
     "file": "HOE00033日韩式餐具合同清单.xlsx", "import_date": "2026-05-14"},
]
for h in hoes:
    es.append(h); existing_ids.add(h['id'])

for i, item in enumerate(items):
    iid = f"HOE_ITEM_JAPAN_{i+1:03d}"
    es.append({
        "id": iid, "type": "hoe_item", "label": item['name'],
        "contract_id": "HOE_CONTRACT_JAPANESE_001",
        "vendor_id": "HOE_VENDOR_JAPANESE_001",
        "brand": item['brand'], "model": item['model'],
        "maker": item['maker'], "unit": item['unit'], "qty": item['qty'],
    })
    existing_ids.add(iid)

fb['entities'] = es
json.dump(fb, open(fb_fp, 'w', encoding='utf-8'), ensure_ascii=False, indent=2)
print(f'\nFB-HOE总实体: {len(es)}')
print(f'+ 2 + {len(items)} = {2+len(items)}')
