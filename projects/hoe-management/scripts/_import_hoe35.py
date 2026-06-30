#!/usr/bin/env python3
import openpyxl, json, os, sys
from collections import Counter
sys.stdout.reconfigure(encoding='utf-8')

BASE = os.path.dirname(os.path.abspath(__file__))
indir = r'C:\Users\Duke Wang\.openclaw\media\inbound'
fn = [f for f in os.listdir(indir) if 'HOE00035' in f or '59beb3d3' in f][0]
fp = os.path.join(indir, fn)

wb = openpyxl.load_workbook(fp, data_only=True)
ws = wb.active

print(f'HOE00035: {os.path.getsize(fp)//1024}KB, sheet: {ws.title}')

# 先读所有行看结构
for r in range(1, min(ws.max_row+1, 15)):
    row = [str(ws.cell(r,c).value or '')[:25] for c in range(1, min(ws.max_column+1, 10))]
    if any(v for v in row if v.strip()):
        print(f'R{r}: {row}')

# 找表头行
header_row = 0
for r in range(1, 10):
    if '名称' in str(ws.cell(r, 2).value or '') or '品名' in str(ws.cell(r, 2).value or ''):
        header_row = r
        break

if header_row == 0:
    # 找不到表头直接扫数据
    items = []
    for r in range(1, ws.max_row+1):
        name = str(ws.cell(r, 1).value or '').strip()
        qty_col = 2
        if name and len(name) > 2 and not any(k in name for k in ['合同','清单','合计','序号','HOE']):
            qty = ws.cell(r, qty_col).value or 0
            try: qty = int(float(str(qty).replace(',','')))
            except: qty = 0
            if qty > 0 or True:
                items.append({'name': name, 'qty': qty, 'spec': str(ws.cell(r, 3).value or '').strip()})
    
    print(f'\n解析到 {len(items)} 项')
    for item in items[:10]:
        print(f'  {item["name"][:20]:20s} | {item["qty"]}件 | {item["spec"][:20]}')
    
    total_qty = sum(item['qty'] for item in items)
    print(f'\n总数量: {total_qty}件')

    # 入库
    fb_fp = os.path.join(BASE, "fb_graph.json")
    fb = json.load(open(fb_fp, 'r', encoding='utf-8'))
    es = fb.get('entities', [])
    existing_ids = set(e.get('id','') for e in es)
    
    for prefix in ['HOE_VENDOR_STEWARD_', 'HOE_CONTRACT_STEWARD_', 'HOE_ITEM_STEWARD_']:
        es[:] = [e for e in es if not e.get('id','').startswith(prefix)]
    existing_ids = set(e.get('id','') for e in es)
    
    hoes = [
        {"id": "HOE_VENDOR_STEWARD_001", "type": "hoe_vendor", "label": "管事部设备供应商",
         "category": "设备器具", "status": "合作中", "import_date": "2026-05-14"},
        {"id": "HOE_CONTRACT_STEWARD_001", "type": "hoe_contract", "label": "HOE00035 管事设备合同清单",
         "vendor_id": "HOE_VENDOR_STEWARD_001", "contract_type": "供应合同",
         "category": "设备器具", "status": "合作中", "items_count": len(items),
         "total_qty": total_qty, "file": "HOE00035管事设备合同清单.xlsx", "import_date": "2026-05-14"},
    ]
    for h in hoes:
        es.append(h); existing_ids.add(h['id'])
    
    for i, item in enumerate(items):
        iid = f"HOE_ITEM_STEWARD_{i+1:03d}"
        es.append({
            "id": iid, "type": "hoe_item", "label": item['name'][:40],
            "contract_id": "HOE_CONTRACT_STEWARD_001",
            "vendor_id": "HOE_VENDOR_STEWARD_001",
            "qty": item['qty'], "spec": item['spec'][:60],
        })
    
    fb['entities'] = es
    json.dump(fb, open(fb_fp, 'w', encoding='utf-8'), ensure_ascii=False, indent=2)
    print(f'\nFB-HOE总实体: {len(es)}')
    print(f'+ 2 + {len(items)} = {2+len(items)}')
