#!/usr/bin/env python3
import openpyxl, json, os, sys
from collections import Counter
sys.stdout.reconfigure(encoding='utf-8')

BASE = os.path.dirname(os.path.abspath(__file__))
indir = r'media/inbound'
fn = [f for f in os.listdir(indir) if 'HOE00035' in f or '59beb3d3' in f][0]
fp = os.path.join(indir, fn)
wb = openpyxl.load_workbook(fp, data_only=True)
ws = wb['合同清单']

items = []
for r in range(4, ws.max_row+1):
    seq = ws.cell(r, 1).value
    if not seq: break
    name = str(ws.cell(r, 2).value or '').strip()
    if not name: continue
    qty_str = str(ws.cell(r, 6).value or '0').strip()
    try: qty = int(float(qty_str))
    except: qty = 0
    items.append({
        'name': name,
        'brand': str(ws.cell(r, 3).value or '').strip(),
        'spec': str(ws.cell(r, 4).value or '').strip(),
        'maker': str(ws.cell(r, 5).value or '').strip(),
        'qty': qty,
        'note': str(ws.cell(r, 8).value or '').strip(),
    })

total_qty = sum(item['qty'] for item in items)
brands = Counter()
for item in items:
    brands[item['brand']] += item['qty']

print(f'管家部设�? {len(items)}�? {total_qty}�?)
print(f'\n品牌:')
for b, q in brands.most_common():
    print(f'  {b:6s}: {q}�?)

# 入库
fb_fp = os.path.join(BASE, "fb_graph.json")
fb = json.load(open(fb_fp, 'r', encoding='utf-8'))
es = fb.get('entities', [])
existing_ids = set(e.get('id','') for e in es)

for prefix in ['HOE_VENDOR_STEWARD_', 'HOE_CONTRACT_STEWARD_', 'HOE_ITEM_STEWARD_']:
    es[:] = [e for e in es if not e.get('id','').startswith(prefix)]
existing_ids = set(e.get('id','') for e in es)

hoes = [
    {"id": "HOE_VENDOR_STEWARD_001", "type": "hoe_vendor", "label": "管家部设备供应商(金宝/勘宝)",
     "category": "清洁用品", "status": "合作�?, "import_date": "2026-05-14"},
    {"id": "HOE_CONTRACT_STEWARD_001", "type": "hoe_contract", "label": "HOE00035 管事设备合同清单",
     "vendor_id": "HOE_VENDOR_STEWARD_001", "contract_type": "供应合同",
     "category": "清洁用品", "status": "合作�?, "items_count": len(items),
     "total_qty": total_qty, "file": "HOE00035管事设备合同清单.xlsx", "import_date": "2026-05-14"},
]
for h in hoes:
    es.append(h); existing_ids.add(h['id'])

for i, item in enumerate(items):
    iid = f"HOE_ITEM_STEWARD_{i+1:03d}"
    es.append({
        "id": iid, "type": "hoe_item", "label": item['name'][:40],
        "contract_id": "HOE_CONTRACT_STEWARD_001",
        "vendor_id": "HOE_VENDOR_STEWARD_001",
        "brand": item['brand'], "spec": item['spec'][:60],
        "maker": item['maker'][:40], "qty": item['qty'],
        "note": item['note'][:40],
    })

fb['entities'] = es
json.dump(fb, open(fb_fp, 'w', encoding='utf-8'), ensure_ascii=False, indent=2)
print(f'\nFB-HOE总实�? {len(es)}')
print(f'+ 2 + {len(items)} = {2+len(items)}')
