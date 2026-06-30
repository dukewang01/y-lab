#!/usr/bin/env python3
import openpyxl, json, os, sys
sys.stdout.reconfigure(encoding='utf-8')

BASE = os.path.dirname(os.path.abspath(__file__))
indir = r'C:\Users\Duke Wang\.openclaw\media\inbound'
fn = [f for f in os.listdir(indir) if 'HOE00033' in f or '1a36370d' in f][0]
fp = os.path.join(indir, fn)
wb = openpyxl.load_workbook(fp, data_only=True)
ws = wb.active

for r in range(1, ws.max_row+1):
    row = []
    for c in range(1, 9):
        v = ws.cell(r,c).value
        row.append(str(v)[:25] if v is not None else '')
    if any(v for v in row if v.strip()):
        print(f'R{r}: {row}')
