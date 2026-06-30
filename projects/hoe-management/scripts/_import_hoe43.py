#!/usr/bin/env python3
import openpyxl, json, os, sys
sys.stdout.reconfigure(encoding='utf-8')

BASE = os.path.dirname(os.path.abspath(__file__))
indir = r'C:\Users\Duke Wang\.openclaw\media\inbound'

fn = [f for f in os.listdir(indir) if 'HOE00043' in f or '3cb19d6f' in f][0]
fp = os.path.join(indir, fn)
wb = openpyxl.load_workbook(fp, data_only=True)
ws = wb['采购清单']

items = []
for r in range(4, ws.max_row+1):
    fixed_id = str(ws.cell(r, 1).value or '').strip()
    seq = ws.cell(r, 2).value
    if not seq: break
    try:
        item = {
            'fixed_asset_id': fixed_id,
            'seq': int(seq),
            'name': str(ws.cell(r, 3).value or '').strip(),
            'brand': str(ws.cell(r, 4).value or '').strip(),
            'model': str(ws.cell(r, 5).value or '').strip(),
            'material': str(ws.cell(r, 6).value or '').strip(),
            'size': str(ws.cell(r, 7).value or '').strip(),
            'unit': str(ws.cell(r, 8).value or '').strip(),
            'qty': int(float(ws.cell(r, 9).value or 0)),
        }
        items.append(item)
    except: pass

total_qty = sum(item['qty'] for item in items)
print(f'西厨器皿: {len(items)} 项, {total_qty} 件')

# 品牌统计
brands = {}
for item in items:
    b = item['brand']
    if b not in brands: brands[b] = 0
    brands[b] += item['qty']
print(f'\n品牌分布:')
for b, q in sorted(brands.items(), key=lambda x: -x[1]):
    print(f'  {b}: {q}件')

# 分类
cats = {}
for item in items:
    n = item['name']
    if '汤桶' in n or '餐炉' in n or '支架' in n: cat = '自助餐设备'
    elif '咖啡' in n or '果汁' in n or '牛奶' in n or '麦片' in n: cat = '饮料分发设备'
    elif '陶瓷内胆' in n: cat = '配件'
    elif '蛋糕' in n: cat = '展示架'
    elif '刀' in n or '叉' in n or '勺' in n or '筷子' in n: cat = '餐具'
    elif '锅' in n or '平底' in n: cat = '厨具'
    elif '保温' in n or '冰' in n: cat = '保温冷藏'
    else: cat = '其他'
    if cat not in cats: cats[cat] = {'count': 0, 'qty': 0}
    cats[cat]['count'] += 1
    cats[cat]['qty'] += item['qty']
print(f'\n分类:')
for c, d in sorted(cats.items(), key=lambda x: -x[1]['qty']):
    print(f'  {c}: {d["count"]}种, {d["qty"]}件')

# 入库FB-HOE
fb_fp = os.path.join(BASE, "fb_graph.json")
fb = json.load(open(fb_fp, 'r', encoding='utf-8'))
es = fb.get('entities', [])
existing_ids = set(e.get('id','') for e in es)

hoes = [
    {"id": "HOE_VENDOR_KITCHEN_001", "type": "hoe_vendor", "label": "西厨设备供应商(Tiger/Vollrath等)",
     "category": "设备器具", "status": "合作中", "import_date": "2026-05-14"},
    {"id": "HOE_CONTRACT_KITCHEN_001", "type": "hoe_contract", "label": "HOE00043 西厨器皿合同",
     "vendor_id": "HOE_VENDOR_KITCHEN_001", "contract_type": "供应合同",
     "category": "设备器具", "status": "合作中", "items_count": len(items),
     "total_qty": total_qty, "file": "HOE00043西厨器皿合同清单.xlsx", "import_date": "2026-05-14"},
]

for h in hoes:
    if h['id'] not in existing_ids:
        es.append(h); existing_ids.add(h['id'])

for item in items:
    iid = f"HOE_ITEM_KITCHEN_{item['seq']:03d}"
    if iid not in existing_ids:
        es.append({
            "id": iid, "type": "hoe_item", "label": item['name'],
            "contract_id": "HOE_CONTRACT_KITCHEN_001",
            "vendor_id": "HOE_VENDOR_KITCHEN_001",
            "brand": item['brand'], "model": item['model'],
            "material": item['material'], "size": item['size'],
            "unit": item['unit'], "qty": item['qty'],
        })
        existing_ids.add(iid)

fb['entities'] = es
json.dump(fb, open(fb_fp, 'w', encoding='utf-8'), ensure_ascii=False, indent=2)
print(f'\nFB-HOE总实体: {len(es)}')
print(f'+ 供应商1 + 合同1 + {len(items)}品项 = {2+len(items)}')
