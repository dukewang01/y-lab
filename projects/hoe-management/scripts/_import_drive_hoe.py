#!/usr/bin/env python3
import openpyxl, json, os, sys, re
sys.stdout.reconfigure(encoding='utf-8')

BASE = os.path.dirname(os.path.abspath(__file__))
D = r'D:\\'

# 找D盘HOE文件
hoe_paths = {}
for root, dirs, files in os.walk(D):
    for f in files:
        m = re.search(r'HOE(\d{5})', f)
        if m and f.endswith(('.xls','.xlsx')):
            hid = m.group(0)
            if hid not in hoe_paths:
                hoe_paths[hid] = os.path.join(root, f)

imported = ['HOE00041','HOE00043','HOE00033','HOE00021','HOE00035',
            'HOE00051','HOE00042','HOE00054','HOE00045','HOE00061','HOE00039']

new_contracts = ['HOE00016','HOE00023','HOE00038','HOE00040','HOE00049','HOE00050','HOE00062']

hoe_config = {
    'HOE00016': {'vendor': '员工餐厅供应商', 'category': '员工餐厅', 'label': '员工餐厅及厨房用具合同'},
    'HOE00023': {'vendor': '清洁药剂供应商', 'category': '清洁用品', 'label': '清洁药剂合同'},
    'HOE00038': {'vendor': '客房房口车供应商', 'category': '设备器具', 'label': '客房房口车合同'},
    'HOE00040': {'vendor': '印刷品供应商', 'category': '印刷物料', 'label': '品牌印刷合同'},
    'HOE00049': {'vendor': '多功能厅设备供应商', 'category': '设备器具', 'label': '多功能厅设备合同'},
    'HOE00050': {'vendor': '布草纺织品供应商', 'category': '布草制服', 'label': '布草纺织品合同'},
    'HOE00062': {'vendor': '意大利餐厅瓷器供应商', 'category': '瓷器', 'label': '意大利餐厅瓷器合同'},
}

# 读取图谱
fb_fp = os.path.join(BASE, "fb_graph.json")
fb = json.load(open(fb_fp, 'r', encoding='utf-8'))
es = fb.get('entities', [])
existing_ids = set(e.get('id','') for e in es)

total_added = 0
new_vendor_ids = []

for hid in new_contracts:
    if hid not in hoe_paths:
        print(f'{hid}: 文件未找到')
        continue
    fp = hoe_paths[hid]
    
    try:
        if fp.endswith('.xls'):
            # 老格式需要特殊处理
            print(f'{hid}: .xls格式跳过（待转换）')
            continue
        
        cfg = hoe_config[hid]
        wb = openpyxl.load_workbook(fp, data_only=True)
        ws = wb.active
        
        # 找表头行
        header_r = 0
        for r in range(1, 8):
            v = str(ws.cell(r, 2).value or ws.cell(r, 3).value or '')
            if '名称' in v or '品名' in v or '产品' in v:
                header_r = r
                break
        start_r = header_r + 1 if header_r else 4
        
        # 解析品项
        items = []
        for r in range(start_r, ws.max_row+1):
            name = str(ws.cell(r, 3).value or ws.cell(r, 2).value or '').strip()
            if not name or len(name) < 2: continue
            if any(k in name for k in ['合计','总价','大写','序号']): continue
            
            qty = 0
            for c in range(6, 11):
                v = ws.cell(r, c).value
                if v and isinstance(v, (int,float)) and 0 < int(v) < 100000:
                    qty = int(v)
                    break
            if qty == 0: continue
            
            brand = str(ws.cell(r, 4).value or '').strip()
            items.append({'name': name, 'brand': brand, 'qty': qty})
        
        if not items:
            print(f'{hid}: 解析到0项，跳过')
            wb.close()
            continue
        
        total_qty = sum(item['qty'] for item in items)
        
        # 入库
        vid = f"HOE_VENDOR_{hid}_001"
        cid = f"HOE_CONTRACT_{hid}_001"
        
        for eid in [vid, cid]:
            es[:] = [e for e in es if e.get('id') != eid]
        for i in range(1, 200):
            es[:] = [e for e in es if e.get('id') != f"HOE_ITEM_{hid}_{i:03d}"]
        existing_ids = set(e.get('id','') for e in es)
        
        new_vendor_ids.append(vid)
        
        hoes = [
            {"id": vid, "type": "hoe_vendor", "label": cfg['vendor'],
             "category": cfg['category'], "status": "合作中"},
            {"id": cid, "type": "hoe_contract", "label": cfg['label'],
             "vendor_id": vid, "items_count": len(items), "total_qty": total_qty},
        ]
        for h in hoes:
            if h['id'] not in existing_ids:
                es.append(h); existing_ids.add(h['id'])
        
        for i, item in enumerate(items):
            iid = f"HOE_ITEM_{hid}_{i+1:03d}"
            if iid not in existing_ids:
                es.append({
                    "id": iid, "type": "hoe_item", "label": item['name'][:40],
                    "contract_id": cid, "vendor_id": vid,
                    "brand": item['brand'], "qty": item['qty'],
                })
                existing_ids.add(iid)
        
        total_added += len(items) + 2
        brands = set(item['brand'] for item in items if item['brand'])
        brand_str = ', '.join(list(brands)[:4]) if brands else '-'
        print(f'{hid}: {len(items)}项, {total_qty}件, 品牌={brand_str}')
        wb.close()
        
    except Exception as e:
        print(f'{hid}: ERROR - {str(e)[:50]}')

# 新增品类节点
new_cats = ['HOE_CATEGORY_EMPLOYEE','HOE_CATEGORY_CHEMICAL','HOE_CATEGORY_LINEN','HOE_CATEGORY_PRINT']
cat_labels = {'HOE_CATEGORY_EMPLOYEE':'员工餐厅','HOE_CATEGORY_CHEMICAL':'清洁药剂','HOE_CATEGORY_LINEN':'布草制服','HOE_CATEGORY_PRINT':'印刷物料'}
for cid, cl in cat_labels.items():
    if cid not in existing_ids:
        es.append({"id": cid, "type": "hoe_category", "label": cl})
        existing_ids.add(cid)

# 版本
for v in [e for e in es if e.get('type')=='version' and 'FIN_VER' in e.get('id','')]:
    es.remove(v)
es.append({'id':'FIN_VER_v5_20_HOE_FULL','type':'version','label':'FIN v5.20 - HOE全酒店体系完成'})

fb['entities'] = es
json.dump(fb, open(fb_fp, 'w', encoding='utf-8'), ensure_ascii=False, indent=2)

print(f'\n=== HOE全酒店体系入库完成 ===')
print(f'新增: {total_added} 实体')
print(f'FB-HOE总实体: {len(es)}')

# 统计
contracts = [e for e in es if e.get('type') == 'hoe_contract']
vendors = [e for e in es if e.get('type') == 'hoe_vendor']
items = [e for e in es if e.get('type') == 'hoe_item']
print(f'\nHOE模块最终:')
print(f'  合同: {len(contracts)}份')
print(f'  供应商: {len(vendors)}家')
print(f'  品项: {len(items)}项')
total_qty = sum(item.get('qty',0) or 0 for item in items)
print(f'  总数量: {total_qty:,}件')
