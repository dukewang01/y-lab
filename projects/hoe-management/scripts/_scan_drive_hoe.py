#!/usr/bin/env python3
import openpyxl, os, sys, re
sys.stdout.reconfigure(encoding='utf-8')

D = r'D:\\'
hoe_files = {}
for root, dirs, files in os.walk(D):
    for f in files:
        m = re.search(r'HOE(\d{5})', f)
        if m and f.endswith(('.xls','.xlsx')):
            hid = m.group(0)
            if hid not in hoe_files:
                hoe_files[hid] = os.path.join(root, f)
    if len(hoe_files) > 25:
        break

# 已入库
imported = ['HOE00041','HOE00043','HOE00033','HOE00021','HOE00035',
            'HOE00051','HOE00042','HOE00054','HOE00045','HOE00061','HOE00039']

new_ones = [hid for hid in sorted(hoe_files.keys()) if hid not in imported]
print(f'D盘HOE合同总数: {len(hoe_files)}')
print(f'已入库: {len(imported)}')
print(f'未入库新合同: {len(new_ones)}个')
print()

for hid in new_ones:
    fp = hoe_files[hid]
    try:
        wb = openpyxl.load_workbook(fp, data_only=True, read_only=True)
        ws = wb.active
        # 读标题
        title = ''
        for r in range(1, 4):
            for c in range(1, 5):
                v = ws.cell(r, c).value
                if v and 'HOE' in str(v):
                    title = str(v).strip()
                    break
            if title: break
        if not title:
            title = str(ws.cell(1,1).value or '')[:40]
        
        # 读品项数
        items = 0
        for r in range(4, min(ws.max_row+1, 20)):
            v = ws.cell(r, 2).value or ws.cell(r, 3).value
            if v and str(v).strip():
                items += 1
        
        # 读品牌
        brands = set()
        for r in range(4, min(ws.max_row+1, 15)):
            b = str(ws.cell(r, 4).value or ws.cell(r, 3).value or '').strip()
            if b and len(b) > 1 and b not in ('序号','名称','品牌'):
                brands.add(b[:15])
        
        sz = os.path.getsize(fp) // 1024
        print(f'{hid:12s} | {sz:>4}KB | 约{items}项 | {title[:30]}')
        if brands:
            b = ', '.join(list(brands)[:4])
            print(f'            品牌: {b}')
        
        wb.close()
    except Exception as e:
        print(f'{hid:12s} | ERROR | {str(e)[:30]}')
    print()
