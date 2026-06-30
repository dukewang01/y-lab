#!/usr/bin/env python3
import openpyxl, os, sys
sys.stdout.reconfigure(encoding='utf-8')

indir = r'C:\Users\Duke Wang\.openclaw\media\inbound'
for fn in os.listdir(indir):
    if 'HOE00041' in fn or 'e1b0624c' in fn:
        fp = os.path.join(indir, fn)
        print(f'Found: {os.path.getsize(fp)//1024}KB')
        wb = openpyxl.load_workbook(fp, data_only=True)
        print(f'Sheets: {wb.sheetnames}')
        for sn in wb.sheetnames:
            ws = wb[sn]
            print(f'\n=== {sn} ({ws.max_row}r x {ws.max_column}c) ===')
            for r in range(1, min(ws.max_row+1, 40)):
                row = []
                for c in range(1, min(ws.max_column+1, 10)):
                    v = ws.cell(r,c).value
                    row.append(str(v)[:30] if v is not None else '')
                if any(v for v in row if v.strip()):
                    print(f'  R{r}: {row}')
        break
