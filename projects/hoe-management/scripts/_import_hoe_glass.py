#!/usr/bin/env python3
import openpyxl, json, os, sys
sys.stdout.reconfigure(encoding='utf-8')

BASE = os.path.dirname(os.path.abspath(__file__))
indir = r'media/inbound'

# 读取Excel
fn = [f for f in os.listdir(indir) if 'HOE00041' in f or 'e1b0624c' in f][0]
fp = os.path.join(indir, fn)
wb = openpyxl.load_workbook(fp, data_only=True)
ws = wb['采购清单']

items = []
for r in range(4, ws.max_row+1):
    seq = ws.cell(r, 1).value
    if not seq: break
    item = {
        'seq': int(seq),
        'name': str(ws.cell(r, 2).value or '').strip(),
        'brand': str(ws.cell(r, 3).value or '').strip(),
        'model': str(ws.cell(r, 4).value or '').strip(),
        'spec': str(ws.cell(r, 5).value or '').strip(),
        'unit': str(ws.cell(r, 6).value or '').strip(),
        'qty': ws.cell(r, 7).value or 0,
        'received': ws.cell(r, 8).value or 0,
        'unit_price': ws.cell(r, 9).value or 0,
    }
    try: item['unit_price'] = float(item['unit_price'])
    except: item['unit_price'] = 0
    try: item['qty'] = int(float(item['qty']))
    except: item['qty'] = 0
    try: item['received'] = int(float(item['received'])) if item['received'] else 0
    except: item['received'] = 0
    items.append(item)

total_qty = sum(item['qty'] for item in items)
total_cost = sum(item['qty'] * item['unit_price'] for item in items)
received_cost = sum((item['received'] or item['qty']) * item['unit_price'] for item in items)

print(f'玻璃器皿合同: {len(items)} �?)
print(f'总订货量: {total_qty} �?)
print(f'合同金额: ¥{total_cost:,.0f}')
print(f'实收金额: ¥{received_cost:,.0f}')

# 按品牌统�?brands = {}
for item in items:
    b = item['brand']
    if b not in brands: brands[b] = {'qty': 0, 'cost': 0}
    brands[b]['qty'] += item['qty']
    brands[b]['cost'] += item['qty'] * item['unit_price']
print(f'\n按品�?')
for b, d in sorted(brands.items(), key=lambda x: -x[1]['cost']):
    print(f'  {b}: {d["qty"]}�? ¥{d["cost"]:,.0f}')

# 入库到FB-HOE
fb_fp = os.path.join(BASE, "fb_graph.json")
fb = json.load(open(fb_fp, 'r', encoding='utf-8'))
es = fb.get('entities', [])
existing_ids = set(e.get('id','') for e in es)

vendor_id = "HOE_VENDOR_GLASS_001"
contract_id = "HOE_CONTRACT_GLASS_001"

# 供应�?hoes = [
    {"id": vendor_id, "type": "hoe_vendor", "label": "玻璃器皿供应�?,
     "contract_id": contract_id, "category": "设备器具", "status": "合作�?,
     "import_date": "2026-05-14", "items_count": len(items)},
    {"id": contract_id, "type": "hoe_contract", "label": "HOE00041 酒店玻璃器皿合同",
     "vendor_id": vendor_id, "contract_type": "供应合同", "category": "设备器具",
     "status": "合作�?, "items_count": len(items), "total_qty": total_qty,
     "total_amount": round(total_cost, 2), "received_amount": round(received_cost, 2),
     "brands": list(brands.keys()), "file": "HOE00041玻璃器皿合同清单.xlsx",
     "import_date": "2026-05-14"},
]

for h in hoes:
    if h['id'] not in existing_ids:
        es.append(h); existing_ids.add(h['id'])

# 每个品项作为hoe_item
for item in items:
    iid = f"HOE_ITEM_GLASS_{item['seq']:03d}"
    if iid not in existing_ids:
        es.append({
            "id": iid, "type": "hoe_item", "label": item['name'],
            "contract_id": contract_id, "vendor_id": vendor_id,
            "brand": item['brand'], "model": item['model'],
            "spec": item['spec'], "unit": item['unit'],
            "qty": item['qty'], "received": item['received'],
            "unit_price": item['unit_price'],
        })
        existing_ids.add(iid)

# 品类节点补充
cat_id = "HOE_CATEGORY_GLASSWARE"
if cat_id not in existing_ids:
    es.append({"id": cat_id, "type": "hoe_category", "label": "玻璃器皿", "description": "酒杯/水杯/餐具�?})

fb['entities'] = es
json.dump(fb, open(fb_fp, 'w', encoding='utf-8'), ensure_ascii=False, indent=2)
print(f'\nFB-HOE总实�? {len(es)}')
print(f'新增: 供应�? + 合同1 + 品类1 + {len(items)}品项 = {3+len(items)}')
