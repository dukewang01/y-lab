#!/usr/bin/env python3
import openpyxl, json, os, sys
sys.stdout.reconfigure(encoding='utf-8')

BASE = os.path.dirname(os.path.abspath(__file__))
indir = r'C:\Users\Duke Wang\.openclaw\media\inbound'

fn = [f for f in os.listdir(indir) if 'HOE00043' in f or '3cb19d6f' in f][0]
fp = os.path.join(indir, fn)
print(f'HOE00043: {os.path.getsize(fp)//1024}KB')

wb = openpyxl.load_workbook(fp, data_only=True)
print(f'Sheets: {wb.sheetnames}')

for sn in wb.sheetnames:
    ws = wb[sn]
    print(f'\n=== {sn} ({ws.max_row}r x {ws.max_column}c) ===')
    for r in range(1, min(ws.max_row+1, 25)):
        row = []
        for c in range(1, min(ws.max_column+1, 10)):
            v = ws.cell(r,c).value
            row.append(str(v)[:30] if v is not None else '')
        if any(v for v in row if v.strip()):
            print(f'  R{r}: {row}')
