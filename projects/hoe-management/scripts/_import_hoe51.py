#!/usr/bin/env python3
import openpyxl, json, os, sys
from collections import Counter
sys.stdout.reconfigure(encoding='utf-8')

BASE = os.path.dirname(os.path.abspath(__file__))
indir = r'C:\Users\Duke Wang\.openclaw\media\inbound'
fn = [f for f in os.listdir(indir) if 'HOE00051' in f or '594cadef' in f][0]
fp = os.path.join(indir, fn)
wb = openpyxl.load_workbook(fp, data_only=True)
ws = wb.active

print(f'HOE00051: {os.path.getsize(fp)//1024}KB')

items = []
for r in range(1, ws.max_row+1):
    name = str(ws.cell(r, 2).value or ws.cell(r, 1).value or '').strip()
    if not name or '名称' in name or '品名' in name or '序号' in name or '合计' in name:
        continue
    # 找数量
    qty = 0
    for c in range(3, min(ws.max_column+1, 8)):
        v = ws.cell(r, c).value
        if v and isinstance(v, (int, float)):
            qty = int(v)
            break
    if qty > 0 or len(name) > 2:
        brand = str(ws.cell(r, 3).value or '').strip()
        spec = str(ws.cell(r, 4).value or '').strip()
        items.append({'name': name, 'brand': brand if brand and brand != name else '', 'spec': spec, 'qty': qty})

# 按行号有表头，找表头行
header_row = 0
ws2 = wb.active
for r in range(1, 8):
    row_vals = [str(ws2.cell(r,c).value or '') for c in range(1, 8)]
    if '名称' in str(row_vals) or '品名' in str(row_vals):
        header_row = r
        break

print(f'表头行: {header_row}')

# 重新正确解析
items2 = []
for r in range(header_row+1 if header_row else 2, ws.max_row+1):
    name = str(ws.cell(r, 2).value or '').strip()
    if not name or len(name) < 2: continue
    if '总价' in name or '合计' in name or '大写' in name: continue
    qty_str = str(ws.cell(r, 6).value or '0').strip()
    try: qty = int(float(qty_str))
    except: qty = 0
    if qty == 0: continue
    items2.append({
        'name': name,
        'brand': str(ws.cell(r, 3).value or '').strip(),
        'spec': str(ws.cell(r, 4).value or '').strip(),
        'maker': str(ws.cell(r, 5).value or '').strip(),
        'qty': qty,
    })

if not items2:
    # fallback: 用第一个扫描
    items2 = items

total_qty = sum(item['qty'] for item in items2)
brands = Counter()
for item in items2:
    if item['brand']: brands[item['brand']] += item['qty']

print(f'西厨厨具: {len(items2)}项, {total_qty}件')
if brands:
    for b, q in brands.most_common(10):
        print(f'  {b:12s}: {q}件')
print(f'\n前5项:')
for item in items2[:5]:
    print(f'  {item["name"][:20]:20s} | {item.get("brand","-"):10s} | {item["qty"]}件')

# 入库
fb_fp = os.path.join(BASE, "fb_graph.json")
fb = json.load(open(fb_fp, 'r', encoding='utf-8'))
es = fb.get('entities', [])
existing_ids = set(e.get('id','') for e in es)

for prefix in ['HOE_VENDOR_WESTKITCHEN_', 'HOE_CONTRACT_WESTKITCHEN_', 'HOE_ITEM_WESTKITCHEN_']:
    es[:] = [e for e in es if not e.get('id','').startswith(prefix)]
existing_ids = set(e.get('id','') for e in es)

hoes = [
    {"id": "HOE_VENDOR_WESTKITCHEN_001", "type": "hoe_vendor", "label": "西厨厨具供应商",
     "category": "设备器具", "status": "合作中", "import_date": "2026-05-14"},
    {"id": "HOE_CONTRACT_WESTKITCHEN_001", "type": "hoe_contract", "label": "HOE00051 西厨厨具合同清单",
     "vendor_id": "HOE_VENDOR_WESTKITCHEN_001", "contract_type": "供应合同",
     "category": "设备器具", "status": "合作中", "items_count": len(items2),
     "total_qty": total_qty, "file": "HOE00051西厨厨具合同清单.xlsx", "import_date": "2026-05-14"},
]
for h in hoes:
    es.append(h); existing_ids.add(h['id'])

for i, item in enumerate(items2):
    iid = f"HOE_ITEM_WESTKITCHEN_{i+1:03d}"
    es.append({
        "id": iid, "type": "hoe_item", "label": item['name'][:40],
        "contract_id": "HOE_CONTRACT_WESTKITCHEN_001",
        "vendor_id": "HOE_VENDOR_WESTKITCHEN_001",
        "brand": item.get('brand',''), "spec": item.get('spec','')[:60],
        "maker": item.get('maker','')[:40], "qty": item['qty'],
    })

fb['entities'] = es
json.dump(fb, open(fb_fp, 'w', encoding='utf-8'), ensure_ascii=False, indent=2)
print(f'\nFB-HOE总实体: {len(es)}')
print(f'+ 2 + {len(items2)} = {2+len(items2)}')
