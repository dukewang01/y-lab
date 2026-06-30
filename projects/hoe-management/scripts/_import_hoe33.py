#!/usr/bin/env python3
import openpyxl, json, os, sys
sys.stdout.reconfigure(encoding='utf-8')

BASE = os.path.dirname(os.path.abspath(__file__))
indir = r'C:\Users\Duke Wang\.openclaw\media\inbound'

fn = [f for f in os.listdir(indir) if 'HOE00033' in f or '1a36370d' in f][0]
fp = os.path.join(indir, fn)
wb = openpyxl.load_workbook(fp, data_only=True)
ws = wb.active

print(f'HOE00033: {os.path.getsize(fp)//1024}KB, sheet: {ws.title}')
print(f'  ({ws.max_row}r x {ws.max_column}c)')

# 找表头
header_row = 0
for r in range(1, 10):
    row_data = []
    for c in range(1, min(ws.max_column+1, 10)):
        v = ws.cell(r,c).value
        row_data.append(str(v)[:20] if v else '')
    print(f'  R{r}: {row_data}')
    if '名称' in str(ws.cell(r,2).value or '') or '品名' in str(ws.cell(r,2).value or ''):
        header_row = r
        break

# 解析数据
items = []
start_row = header_row + 1
for r in range(start_row, ws.max_row+1):
    name = str(ws.cell(r, 2).value or ws.cell(r, 1).value or '').strip()
    if not name or name in ('合计','Total','') or (len(name) < 2 and not any(ws.cell(r,c).value for c in range(1,6))):
        continue
    try:
        item = {
            'name': name,
            'brand': str(ws.cell(r, 3).value or '').strip(),
            'spec': str(ws.cell(r, 4).value or '').strip(),
            'unit': str(ws.cell(r, 5).value or '').strip(),
            'qty': int(float(str(ws.cell(r, 6).value or 0).replace(',',''))) if ws.cell(r,6).value else 0,
        }
        items.append(item)
    except: pass

total_qty = sum(item['qty'] for item in items)
print(f'\n品项: {len(items)}, 总数量: {total_qty}件')
print(f'\n前10项:')
for item in items[:10]:
    print(f'  {item["name"][:20]:20s} | {item["brand"]:10s} | {item["unit"]:4s} | {item["qty"]}件')

# 入库
fb_fp = os.path.join(BASE, "fb_graph.json")
fb = json.load(open(fb_fp, 'r', encoding='utf-8'))
es = fb.get('entities', [])
existing_ids = set(e.get('id','') for e in es)

hoes = [
    {"id": "HOE_VENDOR_JAPANESE_001", "type": "hoe_vendor", "label": "日韩式餐具供应商",
     "category": "设备器具", "status": "合作中", "import_date": "2026-05-14"},
    {"id": "HOE_CONTRACT_JAPANESE_001", "type": "hoe_contract", "label": "HOE00033 日韩式餐具合同",
     "vendor_id": "HOE_VENDOR_JAPANESE_001", "contract_type": "供应合同",
     "category": "设备器具", "status": "合作中", "items_count": len(items),
     "total_qty": total_qty, "file": "HOE00033日韩式餐具合同清单.xlsx", "import_date": "2026-05-14"},
]
for h in hoes:
    if h['id'] not in existing_ids:
        es.append(h); existing_ids.add(h['id'])

for i, item in enumerate(items):
    iid = f"HOE_ITEM_JAPAN_{i+1:03d}"
    if iid not in existing_ids:
        es.append({
            "id": iid, "type": "hoe_item", "label": item['name'],
            "contract_id": "HOE_CONTRACT_JAPANESE_001",
            "vendor_id": "HOE_VENDOR_JAPANESE_001",
            "brand": item['brand'], "spec": item['spec'],
            "unit": item['unit'], "qty": item['qty'],
        })
        existing_ids.add(iid)

fb['entities'] = es
json.dump(fb, open(fb_fp, 'w', encoding='utf-8'), ensure_ascii=False, indent=2)
print(f'\nFB-HOE总实体: {len(es)}')
print(f'+ 供应商1 + 合同1 + {len(items)}品项 = {2+len(items)}')
