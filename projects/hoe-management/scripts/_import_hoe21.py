#!/usr/bin/env python3
import openpyxl, json, os, sys
from collections import Counter
sys.stdout.reconfigure(encoding='utf-8')

BASE = os.path.dirname(os.path.abspath(__file__))
indir = r'media/inbound'
fn = [f for f in os.listdir(indir) if 'HOE00021' in f or '7e7d8a5d' in f][0]
fp = os.path.join(indir, fn)
wb = openpyxl.load_workbook(fp, data_only=True)
ws = wb.active

items = []
for r in range(4, ws.max_row+1):
    seq = ws.cell(r, 1).value
    if not seq and not ws.cell(r, 2).value: continue
    name = str(ws.cell(r, 2).value or '').strip()
    if not name or name == 'åˆè®¡ï¼?: continue
    brand = str(ws.cell(r, 3).value or '').strip()
    spec = str(ws.cell(r, 4).value or '').strip()
    maker = str(ws.cell(r, 5).value or '').strip()
    qty_str = str(ws.cell(r, 6).value or '0').strip()
    try: qty = int(float(qty_str))
    except: qty = 0
    if qty > 0 or name:
        items.append({'name': name, 'brand': brand, 'spec': spec, 'maker': maker, 'qty': qty})

total_qty = sum(item['qty'] for item in items)
print(f'ä¸­å¨æˆ¿åˆå? {len(items)}é¡? {total_qty}ä»?)

brands = Counter()
for item in items:
    brands[item['brand']] += item['qty']
print(f'\nå“ç‰ŒTop 10:')
for b, q in brands.most_common(10):
    print(f'  {b:12s}: {q}ä»?)

makers = Counter()
for item in items:
    makers[item['maker']] += item['qty']
print(f'\nä¾›åº”å•†Top 10:')
for m, q in makers.most_common(10):
    print(f'  {m[:20]:20s}: {q}ä»?)

# åˆ†ç±»
cats = Counter()
for item in items:
    n = item['name']
    if any(k in n for k in ['åˆ€','å‰?,'åˆ?,'åˆ?,'åˆ?,'é”?]): cat = 'åˆ€å…?
    elif any(k in n for k in ['é”?,'é¼?,'å¹³åº•']): cat = 'é”…å…·'
    elif any(k in n for k in ['ç›?,'ç¢?,'ç¢?,'ç›?,'ç›?,'å£?,'æ?]): cat = 'å®¹å™¨'
    elif any(k in n for k in ['ç­?,'æ¼?,'æ»?,'ç½?]): cat = 'ç­›æ¼'
    elif any(k in n for k in ['ç›?,'ç®?,'æ¡?,'ç­?,'ç¯?]): cat = 'å‚¨ç‰©'
    elif any(k in n for k in ['å‹?,'é“?,'åŒ?,'å¤?,'é’?,'å?,'ç­?]): cat = 'å¨å…·'
    elif any(k in n for k in ['æœ?,'å™?,'ç‚?,'æŸ?,'è½?]): cat = 'è®¾å¤‡'
    elif any(k in n for k in ['æ?,'å?,'å¸?,'çº?,'è†?]): cat = 'è€—æ'
    elif any(k in n for k in ['æ?,'æŒ?]): cat = 'æ¶å­'
    elif any(k in n for k in ['æ¨?,'å?]): cat = 'æ¨¡å…·'
    else: cat = 'å…¶ä»–'
    cats[cat] += item['qty']
print(f'\nåˆ†ç±»:')
for c, q in cats.most_common():
    print(f'  {c:6s}: {q}ä»?)

# å…¥åº“
fb_fp = os.path.join(BASE, "fb_graph.json")
fb = json.load(open(fb_fp, 'r', encoding='utf-8'))
es = fb.get('entities', [])
existing_ids = set(e.get('id','') for e in es)

# æ¸…ç†ä¹‹å‰å¯èƒ½å­˜åœ¨çš?for prefix in ['HOE_VENDOR_CHINESE_', 'HOE_CONTRACT_CHINESE_', 'HOE_ITEM_CHINESE_']:
    es[:] = [e for e in es if not e.get('id','').startswith(prefix)]
existing_ids = set(e.get('id','') for e in es)

hoes = [
    {"id": "HOE_VENDOR_CHINESE_001", "type": "hoe_vendor", "label": "ä¸­å¨æˆ¿ç»¼åˆä¾›åº”å•†",
     "category": "è®¾å¤‡å™¨å…·", "status": "åˆä½œä¸?, "import_date": "2026-05-14"},
    {"id": "HOE_CONTRACT_CHINESE_001", "type": "hoe_contract", "label": "HOE00021 ä¸­å¨æˆ¿åˆåŒæ¸…å?,
     "vendor_id": "HOE_VENDOR_CHINESE_001", "contract_type": "ä¾›åº”åˆåŒ",
     "category": "è®¾å¤‡å™¨å…·", "status": "åˆä½œä¸?, "items_count": len(items),
     "total_qty": total_qty, "file": "HOE00021ä¸­å¨æˆ¿åˆåŒæ¸…å?xlsx", "import_date": "2026-05-14"},
]
for h in hoes:
    es.append(h); existing_ids.add(h['id'])

for i, item in enumerate(items):
    iid = f"HOE_ITEM_CHINESE_{i+1:03d}"
    es.append({
        "id": iid, "type": "hoe_item", "label": item['name'][:40],
        "contract_id": "HOE_CONTRACT_CHINESE_001",
        "vendor_id": "HOE_VENDOR_CHINESE_001",
        "brand": item['brand'], "spec": item['spec'][:60],
        "maker": item['maker'][:40], "qty": item['qty'],
    })

fb['entities'] = es
json.dump(fb, open(fb_fp, 'w', encoding='utf-8'), ensure_ascii=False, indent=2)
print(f'\nFB-HOEæ€»å®ä½? {len(es)}')
print(f'+ 2 + {len(items)} = {2+len(items)}')
