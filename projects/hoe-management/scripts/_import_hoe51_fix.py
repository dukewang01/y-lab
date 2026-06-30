#!/usr/bin/env python3
import openpyxl, json, os, sys
from collections import Counter
sys.stdout.reconfigure(encoding='utf-8')

BASE = os.path.dirname(os.path.abspath(__file__))
indir = r'media/inbound'
fn = [f for f in os.listdir(indir) if 'HOE00051' in f or '594cadef' in f][0]
fp = os.path.join(indir, fn)
wb = openpyxl.load_workbook(fp, data_only=True)
ws = wb.active

# 清理错误数量
fb_fp = os.path.join(BASE, "fb_graph.json")
fb = json.load(open(fb_fp, 'r', encoding='utf-8'))
before = len(fb.get('entities',[]))
for prefix in ['HOE_VENDOR_WESTKITCHEN_', 'HOE_CONTRACT_WESTKITCHEN_', 'HOE_ITEM_WESTKITCHEN_']:
    fb['entities'] = [e for e in fb['entities'] if not e.get('id','').startswith(prefix)]
print(f'清理旧数�? {before-len(fb[\"entities\"])}�?)

es = fb['entities']
existing_ids = set(e.get('id','') for e in es)

items = []
for r in range(4, ws.max_row+1):
    name = str(ws.cell(r, 3).value or '').strip()
    if not name or len(name) < 2: continue
    qty = ws.cell(r, 9).value or 0
    try: qty = int(float(str(qty)))
    except: qty = 0
    if qty == 0: continue
    items.append({
        'name': name,
        'brand': str(ws.cell(r, 4).value or '').strip(),
        'material': str(ws.cell(r, 5).value or '').strip(),
        'model': str(ws.cell(r, 6).value or '').strip(),
        'size': str(ws.cell(r, 7).value or '').strip(),
        'qty': qty,
        'location': str(ws.cell(r, 8).value or '').strip(),
    })

total_qty = sum(item['qty'] for item in items)
brands = Counter()
for item in items:
    if item['brand']: brands[item['brand']] += item['qty']

print(f'\n西厨厨具(修正): {len(items)}�? {total_qty}�?)
for b, q in brands.most_common(10):
    print(f'  {b:15s}: {q}�?)

# 入库
hoes = [
    {"id": "HOE_VENDOR_WESTKITCHEN_001", "type": "hoe_vendor", "label": "西厨厨具供应�?Eurochef�?",
     "category": "设备器具", "status": "合作�?, "import_date": "2026-05-14"},
    {"id": "HOE_CONTRACT_WESTKITCHEN_001", "type": "hoe_contract", "label": "HOE00051 西厨厨具合同清单",
     "vendor_id": "HOE_VENDOR_WESTKITCHEN_001", "contract_type": "供应合同",
     "category": "设备器具", "status": "合作�?, "items_count": len(items),
     "total_qty": total_qty, "file": "HOE00051西厨厨具合同清单.xlsx", "import_date": "2026-05-14"},
]
for h in hoes:
    es.append(h); existing_ids.add(h['id'])

for i, item in enumerate(items):
    iid = f"HOE_ITEM_WESTKITCHEN_{i+1:03d}"
    es.append({
        "id": iid, "type": "hoe_item", "label": item['name'][:40],
        "contract_id": "HOE_CONTRACT_WESTKITCHEN_001",
        "vendor_id": "HOE_VENDOR_WESTKITCHEN_001",
        "brand": item['brand'], "model": item['model'][:40],
        "material": item['material'][:20], "size": item['size'][:40],
        "location": item['location'][:20], "qty": item['qty'],
    })

fb['entities'] = es
json.dump(fb, open(fb_fp, 'w', encoding='utf-8'), ensure_ascii=False, indent=2)
print(f'\nFB-HOE总实�? {len(es)}')
print(f'+ 2 + {len(items)} = {2+len(items)}')
