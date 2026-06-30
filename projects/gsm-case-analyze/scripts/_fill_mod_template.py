#!/usr/bin/env python3
import openpyxl, os, shutil, sys
sys.stdout.reconfigure(encoding='utf-8')

src = r'C:\Users\Duke Wang\.openclaw\media\inbound\MOD_report_20260510---338b19d2-5f04-4559-b3dc-71a3c1cdc30a.xlsx'
out = os.path.expanduser(r'~/.openclaw/media/outbound/MOD_Report_20260514.xlsx')
shutil.copy2(src, out)

wb = openpyxl.load_workbook(out)

# ====== Sheet 1: Report1-Guest Area ======
ws = wb['Report1-Guest Area']
# 修改日期和MOD名字
ws.cell(3, 1).value = 'MOD Name 值班主任: Duke Wang'
ws.cell(3, 2).value = 'Date日期: 05/14'

# 1. Entrance/Lobby - 19:00
ws.cell(5, 2).value = 'InspectionTime 检查时间：19:00'
ws.cell(5, 3).value = ''  # 清空旧数据
# 大门锁异响 - 这一行找对位置
row = 6
ws.cell(row, 2).value = ''  # YES空
ws.cell(row, 3).value = 'Y'  # NO
ws.cell(row, 4).value = '大门锁具异响，已报工程部检修'

# 其他检查项正常填YES
for r in range(7, 14):
    ws.cell(r, 2).value = 'Y'
    ws.cell(r, 3).value = ''

# 2. YUAN Lounge 大堂吧 - 收入1500
ws.cell(15, 2).value = 'InspectionTime 检查时间：19:30'
for r in range(16, 24):
    ws.cell(r, 2).value = 'Y'
    ws.cell(r, 3).value = ''
ws.cell(24, 4).value = '当日收入¥1,500'  # 收入行

# 3. Open Restaurant
ws.cell(25, 2).value = 'InspectionTime 检查时间：20:00'
for r in range(26, 33):
    ws.cell(r, 2).value = 'Y'
    ws.cell(r, 3).value = ''
# covers行
for r in range(30, 35):
    v = ws.cell(r, 4).value
    if v and 'covers' in str(v).lower():
        ws.cell(r, 4).value = '晚餐约50人'
        break

# 4. YUXI 御玺
# 找YUXI行
yuxi_row = None
for r in range(35, 60):
    v = str(ws.cell(r, 1).value or '')
    if 'YUXI' in v or '御玺' in v:
        yuxi_row = r
        break
if yuxi_row:
    ws.cell(yuxi_row, 2).value = 'InspectionTime 检查时间：20:30'
    for r2 in range(yuxi_row+1, yuxi_row+8):
        ws.cell(r2, 2).value = 'Y'
        ws.cell(r2, 3).value = ''
    ws.cell(yuxi_row+2, 4).value = '3PDR团队用餐'

# 5. 42F Executive Lounge
lounge_row = None
for r in range(60, 90):
    v = str(ws.cell(r, 1).value or '') + str(ws.cell(r, 2).value or '')
    if 'Lounge' in v or '酒廊' in v:
        lounge_row = r
        break
if lounge_row:
    ws.cell(lounge_row, 2).value = 'InspectionTime 检查时间：21:00'
    for r2 in range(lounge_row+1, lounge_row+8):
        ws.cell(r2, 2).value = 'Y'
        ws.cell(r2, 3).value = ''
    ws.cell(lounge_row+4, 4).value = '欢乐时光约30人'

# 6. Fitness & Pool
fit_row = None
for r in range(90, 130):
    v = str(ws.cell(r, 1).value or '') + str(ws.cell(r, 2).value or '')
    if 'Fitness' in v or '健身' in v or 'Pool' in v or '泳池' in v:
        fit_row = r
        break
if fit_row:
    ws.cell(fit_row, 2).value = 'InspectionTime 检查时间：21:30'
    for r2 in range(fit_row+1, fit_row+12):
        ws.cell(r2, 2).value = 'Y'
        ws.cell(r2, 3).value = ''
    # 泳池教学
    ws.cell(fit_row+3, 4).value = '泳池：8人教学课'
    # 健身房拖鞋
    ws.cell(fit_row+7, 3).value = 'Y'
    ws.cell(fit_row+7, 4).value = '⚠️ 有客人穿酒店拖鞋使用器材，已现场提醒并通知健身中心加强巡查'

# 7. 6F Banquet
bqt_row = None
for r in range(130, 170):
    v = str(ws.cell(r, 1).value or '') + str(ws.cell(r, 2).value or '')
    if 'Banquet' in v or '宴会' in v:
        bqt_row = r
        break
if bqt_row:
    ws.cell(bqt_row, 2).value = 'InspectionTime 检查时间：22:00'
    for r2 in range(bqt_row+1, bqt_row+8):
        ws.cell(r2, 2).value = 'Y'
        ws.cell(r2, 3).value = ''
    ws.cell(bqt_row+2, 4).value = '会议进行中'

# ====== Sheet 2: Report2-Back of House ======
ws2 = wb['Report2-Back of House']
ws2.cell(3, 1).value = 'MOD Name 值班主任: duke'
ws2.cell(3, 2).value = 'Date日期: 05/14'
# 后区正常，全部YES
row = 5
while row <= ws2.max_row:
    v = str(ws2.cell(row, 1).value or '')
    if 'InspectionTime' in v:
        ws2.cell(row, 2).value = v.replace('20:00','22:30').replace('22:00','22:30').replace('16:30','22:30')
        # 检查时间已存在
    elif str(ws2.cell(row, 2).value or '') in ('', 'YES', 'NO'):
        ws2.cell(row, 2).value = 'Y'
        ws2.cell(row, 3).value = ''
    row += 1
ws2.cell(3, 3).value = ''  # 清NO列

# ====== Sheet 3: reports-Rooms ======
ws3 = wb['reports-Rooms']
ws3.cell(2, 5).value = 'MOD Name 值班主任: duke'
ws3.cell(2, 7).value = 'Date日期: 05月14日'
ws3.cell(3, 3).value = 'Guest Room No:3025'
ws3.cell(3, 6).value = 'Guest Room No:4018'
ws3.cell(5, 3).value = '23:00:00'

# 3025 - 异味问题
# Door Lock
ws3.cell(7, 3).value = 'Y'
ws3.cell(7, 4).value = ''
ws3.cell(7, 6).value = 'Y'
ws3.cell(7, 7).value = ''
ws3.cell(7, 8).value = ''

# 逐个检查项 - 3025正常, 4018部分异常
# 用循环填大部分为Y
for r in range(7, 65):
    if ws3.cell(r, 3).value not in ('Y', 'N', ''):
        continue
    # 3025列
    ws3.cell(r, 3).value = 'Y'
    ws3.cell(r, 4).value = ''
    # 4018列
    ws3.cell(r, 6).value = 'Y'
    ws3.cell(r, 7).value = ''
    ws3.cell(r, 8).value = ''

# 3025 - 异味在第14行后备注
sense_row = None
for r in range(50, 80):
    v = str(ws3.cell(r, 1).value or '').lower()
    if 'smell' in v or 'odor' in v or '异味' in v:
        sense_row = r
        break
if not sense_row:
    for r in range(50, 80):
        v = str(ws3.cell(r, 2).value or '').lower()
        if '空气' in v or '通风' in v:
            sense_row = r
            break
if sense_row:
    ws3.cell(sense_row, 5).value = '房间有异味'
    ws3.cell(sense_row, 8).value = ''

# 4018 - 毛发
hair_row = None
for r in range(60, 90):
    v = str(ws3.cell(r, 2).value or '').lower()
    if 'pillow' in v or '枕头' in v or '床单' in v or 'bed' in v:
        hair_row = r
        break
if not hair_row:
    hair_row = 30  # 没有具体匹配行，用30行左右写备注
# 在Remarks列写备注
ws3.cell(15, 8).value = '床头有毛发'
ws3.cell(16, 8).value = ''
ws3.cell(17, 8).value = ''

# 在底部的通用备注区域添加
for r in range(120, 150):
    v = str(ws3.cell(r, 1).value or '')
    if 'Remarks' in v or '备注' in v or 'comment' in v:
        ws3.cell(r, 2).value = '3025房间异味 | 4018床头有毛发'
        break

# ====== Sheet 4: FoodTasting ======
# 今天没有食品试餐，留空

# ====== Sheet 5: Safety & Security ======
ws5 = wb['Safety & Security Checklist']
ws5.cell(2, 2).value = 'MOD: duke'
ws5.cell(2, 3).value = 'Date: 05/14'
for r in range(4, ws5.max_row+1):
    c2 = str(ws5.cell(r, 2).value or '')
    # 所有检查项填Y（正常）
    if c2 and c2 not in ('Question 问题', 'Y/N 是/否'):
        ws5.cell(r, 3).value = 'Y'
        ws5.cell(r, 4).value = ''

wb.save(out)
sz = os.path.getsize(out)
print(f'DONE! {sz//1024} KB')
print(f'Saved: {out}')
